"""Glossary update from post-editing results.

The project glossary is an aligned asset pair (``zh_asset.json`` /
``en_asset.json`` keyed ``asset_NNNN``) plus ``dedup_index.json`` mapping
each asset to its PO msgctxt keys. After a post-editing round the EN side
is stale; the ground truth is scattered across:

- the PE'd bilingual .po (msgid=zh, msgstr=final EN) — per-string truth;
- a decisions workbook (映射留档 / 翻译裁定 sheets) — explicit term
  rulings, but only for PART of what changed;
- extra term pairs the operator supplies for rulings the workbook missed.

``refresh_glossary`` is deterministic and evidence-first:

1. Join every asset to its PE renderings via the dedup index (never by
   text similarity).
2. One agreeing rendering → update. Conflicting renderings (PE updated
   only some duplicate copies) → resolve by term-decision compliance,
   then by majority; still ambiguous → keep old, flag for review.
3. Term audit: every asset whose zh contains a decided term but whose
   final EN lacks the decided rendering is flagged — the glossary is law,
   so violations are surfaced, never silently accepted.
4. Mining: aggregate the word-level replacements across all updated
   entries and report frequent old→new EN phrase pairs that are NOT yet
   covered by any decision — the "nobody wrote it down" terms.

Source files are read-only; updated JSON goes to the output directory in
the same serialization style (indent 2, key order preserved).
"""
from __future__ import annotations

import difflib
import json
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from .exports import read_po_entries
from .po_compare import _WORD_TOKEN_RE
from .standards import FORM_DROPDOWNS, GLOSSARY_PE_FORM_HEADERS

MAPPING_SHEET = "内部·映射留档"
RULING_SHEET = "内部·翻译裁定"


@dataclass
class TermDecision:
    zh: str                       # current source term (dev final)
    en: str                       # decided EN rendering
    zh_old: Optional[str] = None  # superseded source term, if renamed
    origin: str = "user"          # M-05 / J-03 / user
    family: bool = False          # 族 ruling: governs every term
                                  # CONTAINING zh — a RULE, never a
                                  # glossary entry of its own


@dataclass
class GlossaryUpdate:
    updated_en: Dict[str, str]
    decisions: List[TermDecision]
    updated: List[dict] = field(default_factory=list)
    conflicts_resolved: List[dict] = field(default_factory=list)
    conflicts_open: List[dict] = field(default_factory=list)
    unresolved: List[str] = field(default_factory=list)   # asset ids
    unchanged: int = 0
    term_violations: List[dict] = field(default_factory=list)
    suggestions: List[dict] = field(default_factory=list)

    def counts(self) -> Dict[str, int]:
        return {"updated": len(self.updated),
                "conflicts_resolved": len(self.conflicts_resolved),
                "conflicts_open": len(self.conflicts_open),
                "unresolved": len(self.unresolved),
                "unchanged": self.unchanged,
                "term_violations": len(self.term_violations),
                "suggested_terms": len(self.suggestions)}


def _clean(value: object) -> str:
    return str(value).strip() if value is not None else ""


def load_decisions_xlsx(path: Path) -> List[TermDecision]:
    """Explicit term rulings from the decisions workbook.

    映射留档 rows carry (旧源文, 新源文, EN定稿) directly. 翻译裁定 rows
    encode the term in the 议题 column (``铁匠铺EN``) and the ruling as
    ``B → Blacksmith``; rows that do not parse are skipped — this loader
    only trusts unambiguous rows, everything else stays operator work."""
    import openpyxl
    book = openpyxl.load_workbook(path, data_only=True)
    decisions: List[TermDecision] = []

    if MAPPING_SHEET in book.sheetnames:
        rows = list(book[MAPPING_SHEET].iter_rows(values_only=True))
        header = [_clean(h) for h in rows[0]]
        idx = {name: i for i, name in enumerate(header)}
        def col(row, prefix):
            for name, i in idx.items():
                if name.startswith(prefix):
                    return _clean(row[i])
            return ""
        for row in rows[1:]:
            zh_new, en = col(row, "新源文"), col(row, "EN定稿")
            if zh_new and en:
                decisions.append(TermDecision(
                    zh=zh_new, en=en, zh_old=col(row, "旧源文") or None,
                    origin=col(row, "映射ID") or "mapping"))

    if RULING_SHEET in book.sheetnames:
        rows = list(book[RULING_SHEET].iter_rows(values_only=True))
        header = [_clean(h) for h in rows[0]]
        idx = {name: i for i, name in enumerate(header)}
        for row in rows[1:]:
            topic = _clean(row[idx.get("议题", 1)])
            verdict = _clean(row[idx.get("裁定结果(填这里)", 5)])
            match = re.match(r"^([一-鿿/]+?)EN\b", topic)
            arrow = re.search(r"→\s*(.+)$", verdict)
            if not (match and arrow):
                continue
            en_parts = [p.strip() for p in arrow.group(1).split("/")]
            zh_parts = match.group(1).split("/")
            rid = _clean(row[idx.get("裁定ID", 0)]) or "ruling"
            for i, zh in enumerate(zh_parts):
                en = en_parts[min(i, len(en_parts) - 1)]
                # FAMILY rulings: ``致幻族EN … A → Hallucination 族``
                # rules the whole 致幻 term family, the trailing 族 is
                # ruling shorthand ("across the family"), never part of
                # the term or its EN. Family members stay separate
                # glossary entries; the ruling only validates them.
                family = zh.endswith("族") and en.endswith("族")
                if family:
                    zh = zh[:-1]
                    en = en[:-1].strip()
                decisions.append(TermDecision(
                    zh=zh, en=en, family=family,
                    origin=rid + (" (族/family)" if family else "")))
    return decisions


def _en_base(rendering: str) -> str:
    """Workbook finals use gloss notation like ``Apostle(s)`` — match on
    the base form so both singular and plural satisfy the ruling."""
    return re.sub(r"\(s\)", "", rendering, flags=re.I).strip()


def _en_has(rendering: str, text: str) -> bool:
    return _en_base(rendering).lower() in text.lower()


def _applicable(decisions: List[TermDecision], zh_text: str
                ) -> List[TermDecision]:
    return [d for d in decisions
            if d.zh in zh_text or (d.zh_old and d.zh_old in zh_text)]


def refresh_glossary(assets_dir: Path, pe_po: Path,
                     decisions: List[TermDecision]) -> GlossaryUpdate:
    assets_dir = Path(assets_dir)
    zh = json.loads((assets_dir / "zh_asset.json").read_text("utf-8"))
    en = json.loads((assets_dir / "en_asset.json").read_text("utf-8"))
    dedup = json.loads(
        (assets_dir / "dedup_index.json").read_text("utf-8"))["asset"]
    by_key: Dict[str, str] = {
        key.lstrip(","): target
        for key, _source, target, _loc in read_po_entries(Path(pe_po))
        if target.strip()}

    result = GlossaryUpdate(updated_en=dict(en), decisions=decisions)
    for aid, old_en in en.items():
        keys = dedup.get(aid, {}).get("keys", [])
        renderings = [by_key[k] for k in keys if k in by_key]
        if not renderings:
            result.unresolved.append(aid)
            continue
        counts = Counter(renderings)
        entry = {"asset": aid, "zh": zh.get(aid, ""), "old_en": old_en}
        if len(counts) == 1:
            new_en = renderings[0]
            if new_en == old_en:
                result.unchanged += 1
            else:
                result.updated_en[aid] = new_en
                result.updated.append({**entry, "new_en": new_en})
            continue
        # conflicting PE renderings (duplicates updated unevenly):
        # decision compliance first, majority second, else review
        applicable = _applicable(decisions, zh.get(aid, ""))
        compliant = [r for r in counts
                     if applicable
                     and all(_en_has(d.en, r) for d in applicable)]
        if len(compliant) == 1:
            chosen, how = compliant[0], "decision"
        else:
            (top, top_n), *rest = counts.most_common(2)
            if not rest or top_n > rest[0][1]:
                chosen, how = top, "majority"
            else:
                result.conflicts_open.append(
                    {**entry, "renderings": sorted(counts)})
                continue
        if chosen != old_en:
            result.updated_en[aid] = chosen
            result.conflicts_resolved.append(
                {**entry, "new_en": chosen, "resolved_by": how,
                 "renderings": sorted(counts)})
        else:
            result.unchanged += 1

    # term audit over the FINAL state (deduped: two workbook rows ruling
    # the same term must not double-report an asset)
    seen: set = set()
    for decision in decisions:
        for aid, zh_text in zh.items():
            if decision.zh in zh_text or (decision.zh_old
                                          and decision.zh_old in zh_text):
                final = result.updated_en[aid]
                mark = (aid, decision.zh, _en_base(decision.en).lower())
                if not _en_has(decision.en, final) and mark not in seen:
                    seen.add(mark)
                    result.term_violations.append({
                        "asset": aid, "term": decision.zh,
                        "expected_en": decision.en,
                        "origin": decision.origin,
                        "zh": zh_text[:60], "en": final[:80]})

    result.suggestions = _mine_suggestions(
        result.updated + result.conflicts_resolved, decisions)
    return result


def _mine_suggestions(changed: List[dict],
                      decisions: List[TermDecision],
                      min_count: int = 3) -> List[dict]:
    """Frequent old→new EN phrase replacements across updated entries
    that no decision covers — candidate glossary rulings nobody wrote
    down. Evidence counting only; a human promotes them."""
    known = {_en_base(d.en).lower() for d in decisions if d.en.strip()}
    known_words = {word for base in known for word in base.split()}

    def covered(phrase: str) -> bool:
        """Already ruled: exact/containment match on a decided base, or
        every word (singular-normalized) belongs to decided bases —
        catches inflection echoes like 'apostle → apostles'."""
        words = phrase.split()
        return (any(k in phrase or phrase in k for k in known)
                or bool(words) and all(w in known_words
                                       or w.rstrip("s") in known_words
                                       for w in words))

    pairs: Counter = Counter()
    for entry in changed:
        a = _WORD_TOKEN_RE.findall(entry["old_en"].lower())
        b = _WORD_TOKEN_RE.findall(entry["new_en"].lower())
        matcher = difflib.SequenceMatcher(a=a, b=b, autojunk=False)
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == "replace" and i2 - i1 <= 4 and j2 - j1 <= 4:
                pairs[(" ".join(a[i1:i2]), " ".join(b[j1:j2]))] += 1
    return [{"old_en": old, "new_en": new, "occurrences": count}
            for (old, new), count in pairs.most_common()
            if count >= min_count and not covered(new)]


# ------------------------------------------------------- term distillation

TERM_MAX_ZH_CHARS = 10
TERM_MAX_EN_WORDS = 4
_SENTENCE_CHARS = set("。，！？；：、…\r\n的")


def is_term_like(zh: str, en: str) -> bool:
    """A glossary TERM, not a sentence or a templated UI string: short
    zh that actually contains CJK, no sentence punctuation, no
    descriptive 的-phrases, not digit/bracket-led (``1号玩家``, ``(0/3)``,
    ``%`` are UI patterns, not terminology), short EN."""
    from .po_sanity import CJK_RE
    zh = zh.strip()
    return (0 < len(zh) <= TERM_MAX_ZH_CHARS
            and bool(CJK_RE.search(zh))
            and not re.match(r"[\d(（%<\[]", zh)
            and not (set(zh) & _SENTENCE_CHARS)
            and bool(en.strip())
            and len(en.split()) <= TERM_MAX_EN_WORDS)


def distill_term_glossary(zh_asset: Dict[str, str],
                          en_final: Dict[str, str],
                          decisions: List[TermDecision],
                          *, game: str = "", locale: str = "en"
                          ) -> Tuple[dict, List[dict]]:
    """Compact term-level glossary from the full asset pair + rulings.

    Decisions are locked and always win. Term-like assets contribute
    mined entries; identical zh appearing with different EN resolves by
    majority, ties go to the review list instead of the glossary. Output
    is the pipeline T1 file shape ({"metadata", "terms"}), so the result
    plugs straight into Glossary.load_t1_file / LQA locked-term checks."""
    votes: Dict[str, Counter] = {}
    for aid, zh_text in zh_asset.items():
        en_text = en_final.get(aid, "")
        if is_term_like(zh_text, en_text):
            votes.setdefault(zh_text.strip(), Counter())[
                en_text.strip()] += 1

    terms: Dict[str, dict] = {}
    review: List[dict] = []
    for zh_term, counter in sorted(votes.items()):
        (top, top_n), *rest = counter.most_common(2)
        if rest and rest[0][1] == top_n:
            review.append({"zh": zh_term,
                           "renderings": sorted(counter),
                           "reason": "tie between PE renderings"})
            continue
        terms[zh_term] = {"translation": top, "type": "mined",
                          "locked": False,
                          "evidence": f"PE ×{sum(counter.values())}"}
    for decision in decisions:
        if decision.family:            # rules validate, never enter terms
            continue
        terms[decision.zh] = {
            "translation": _en_base(decision.en), "type": "decision",
            "locked": True, "evidence": decision.origin,
            **({"supersedes": decision.zh_old}
               if decision.zh_old else {})}

    # internal consistency: a mined term that CONTAINS a locked term (or
    # falls under a 族 family rule) must render it — 'Ai玩家太刀 =
    # AI Player Greatsword' violates 太刀=Tachi
    flagged = 0
    for zh_term, entry in terms.items():
        if entry["locked"]:
            continue
        for decision in decisions:
            contains = (decision.zh in zh_term if decision.family
                        else decision.zh != zh_term
                        and decision.zh in zh_term)
            if contains and not _en_has(decision.en,
                                        entry["translation"]):
                kind = "family rule" if decision.family else "locked"
                entry["check"] = (f"contains {kind} {decision.zh}="
                                  f"{_en_base(decision.en)}")
                flagged += 1
                break

    family_rules = {d.zh: {"translation": _en_base(d.en),
                           "origin": d.origin}
                    for d in decisions if d.family}
    glossary = {
        "metadata": {"game": game, "locale": locale,
                     "built_from": "PE-refreshed assets + decisions",
                     "locked_terms": sum(1 for t in terms.values()
                                         if t["locked"]),
                     "mined_terms": sum(1 for t in terms.values()
                                        if not t["locked"]),
                     "flagged_against_locked": flagged,
                     **({"family_rules": family_rules}
                        if family_rules else {})},
        "terms": dict(sorted(terms.items(),
                             key=lambda kv: (not kv[1]["locked"],
                                             kv[0])))}
    return glossary, review


def write_term_glossary(glossary: dict, review: List[dict],
                        out_dir: Path) -> Path:
    """glossary_terms.json (pipeline T1 shape) + a compact xlsx list."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "glossary_terms.json").write_text(
        json.dumps(glossary, ensure_ascii=False, indent=2),
        encoding="utf-8")

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Glossary"
    headers = ["zh", "EN", "Locked", "Source", "Notes"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="333F50")
    ordered = sorted(
        glossary["terms"].items(),
        key=lambda kv: (not kv[1]["locked"], "check" not in kv[1], kv[0]))
    for zh_term, entry in ordered:
        notes = " ; ".join(filter(None, (
            "supersedes " + entry["supersedes"]
            if entry.get("supersedes") else "",
            "⚠ " + entry["check"] if entry.get("check") else "")))
        sheet.append([zh_term, entry["translation"],
                      "LOCKED" if entry["locked"] else "",
                      entry["evidence"], notes])
    for row in sheet.iter_rows(min_row=2):
        if row[2].value == "LOCKED":
            for cell in row:
                cell.fill = PatternFill("solid", start_color="FFF2CC")
        elif str(row[4].value or "").startswith("⚠"):
            for cell in row:
                cell.fill = PatternFill("solid", start_color="FFC7A0")
    families = glossary.get("metadata", {}).get("family_rules") or {}
    if families:
        fam_sheet = book.create_sheet("Family Rules 术语族")
        fam_sheet.append(["zh family base", "EN rendering", "Origin",
                          "Rule"])
        for zh_base, rule in families.items():
            fam_sheet.append([zh_base, rule["translation"],
                              rule.get("origin", ""),
                              f"every term containing {zh_base} must "
                              f"render it as {rule['translation']!r}"])
        for letter, width in (("A", 18), ("B", 22), ("C", 20), ("D", 55)):
            fam_sheet.column_dimensions[letter].width = width
    if review:
        tie_sheet = book.create_sheet("Ties (excluded)")
        tie_sheet.append(["zh", "Renderings", "Reason"])
        for entry in review:
            tie_sheet.append([entry["zh"],
                              " | ".join(entry["renderings"]),
                              entry["reason"]])
        tie_sheet.column_dimensions["A"].width = 30
        tie_sheet.column_dimensions["B"].width = 60
    for letter, width in (("A", 28), ("B", 36), ("C", 10), ("D", 22),
                          ("E", 28)):
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"
    path = out_dir / "glossary_terms.xlsx"
    book.save(path)
    return path


TYPE_FILL = {"Conflict": "FFC7A0", "Violation": "FF9999",
             "Suggestion": "D9E1F2", "Update": "E2EFDA"}


def write_review_xlsx(result: GlossaryUpdate, path: Path) -> int:
    """Glossary PE form (standards.GLOSSARY_PE_FORM_HEADERS) — same
    family as the MTPE/LQA PE forms: Target_Original/Target_Suggested
    pair, PE_* block, LQA decision vocabulary. One row per item,
    review-first ordering (open conflicts → term violations → mined
    suggestions → auto-updates). Returns the row count."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    headers = list(GLOSSARY_PE_FORM_HEADERS)
    rows: List[dict] = []
    for e in result.conflicts_open:
        rows.append({
            "TermID": e["asset"], "EntryType": "Conflict",
            "Source": e["zh"], "Target_Original": e["old_en"],
            "Alternatives": " | ".join(e["renderings"]),
            "Evidence": "PE renderings disagree; no ruling/majority — "
                        "pick one via Reject&Modification",
        })
    for e in result.term_violations:
        rows.append({
            "TermID": e["asset"], "EntryType": "Violation",
            "Source": e["zh"], "Target_Original": e["en"],
            "Target_Suggested": f"(must render {e['term']} as "
                                f"{e['expected_en']!r})",
            "Evidence": e["origin"],
        })
    for s in result.suggestions:
        rows.append({
            "EntryType": "Suggestion", "Target_Original": s["old_en"],
            "Target_Suggested": s["new_en"],
            "Evidence": f"mined from PE, ×{s['occurrences']}",
        })
    for e in result.updated + result.conflicts_resolved:
        rows.append({
            "TermID": e["asset"], "EntryType": "Update",
            "Source": e["zh"], "Target_Original": e["old_en"],
            "Target_Suggested": e["new_en"],
            "Alternatives": " | ".join(e.get("renderings", [])),
            "Evidence": "PE (" + e.get("resolved_by", "agreed") + ")",
        })

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Glossary PE"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="333F50")
        cell.alignment = Alignment(vertical="center")
    for number, row in enumerate(rows, 1):
        sheet.append([row.get(h, "") for h in headers])
        fill = PatternFill("solid",
                           start_color=TYPE_FILL[row["EntryType"]])
        sheet.cell(row=number + 1,
                   column=headers.index("EntryType") + 1).fill = fill
    for col, options in FORM_DROPDOWNS["glossary"].items():
        validation = DataValidation(
            type="list", formula1='"' + ",".join(options) + '"',
            allow_blank=True, showDropDown=False)
        sheet.add_data_validation(validation)
        letter = sheet.cell(
            row=1, column=headers.index(col) + 1).column_letter
        validation.add(f"{letter}2:{letter}{len(rows) + 1}")
    for name, width in (("Source", 40), ("Target_Original", 45),
                        ("Target_Suggested", 45), ("Alternatives", 45),
                        ("Evidence", 30), ("PE_Decision", 28),
                        ("PE_Modification", 45), ("PE_Note", 25),
                        ("PE_Query", 25)):
        letter = sheet.cell(
            row=1, column=headers.index(name) + 1).column_letter
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"

    guide = book.create_sheet("说明 How to fill")
    for line in (
            "One row per glossary item needing (or documenting) a change.",
            "Same form family as the MTPE / LQA PE tables: fill only the "
            "PE_* columns; agent columns must come back unchanged.",
            "",
            "PE_Decision (dropdown):",
            "  Accept Suggested Translation — apply Target_Suggested "
            "as-is. Conflict rows carry no suggestion: use "
            "Reject&Modification with your pick.",
            "  Reject&Modification — your wording wins: final EN in "
            "PE_Modification, how/why in PE_Note.",
            "  Reject&Keep-as-it-is — keep Target_Original (on Update "
            "rows this rolls back the auto-applied change).",
            "  Reject&Cannot Answer — explain in PE_Query; it is "
            "forwarded to the dev team and the row stays open.",
            "  (blank) — undecided; the row is carried to the next round.",
            "",
            "PE_Categorization (Accuracy/Terminology/Tone/Fluency/"
            "Technical) and PE_Severity (Critical/Major/Minor) are "
            "optional classification.",
            "",
            "EntryType: Conflict = PE duplicates disagree (orange) · "
            "Violation = decided term missing (red) · Suggestion = mined "
            "ruling candidate (blue) · Update = auto-applied from PE, "
            "review optional (green)."):
        guide.append([line])
    guide.column_dimensions["A"].width = 110
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return len(rows)


def write_update_outputs(result: GlossaryUpdate, out_dir: Path,
                         assets_dir: Path) -> Path:
    """Updated en_asset.json (same style, key order preserved) +
    termbase delta + audit report. Originals are never touched."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "en_asset.json").write_text(
        json.dumps(result.updated_en, ensure_ascii=False, indent=2),
        encoding="utf-8")
    (out_dir / "termbase_delta.json").write_text(json.dumps(
        [{"zh": d.zh, "zh_old": d.zh_old, "en": d.en, "origin": d.origin}
         for d in result.decisions], ensure_ascii=False, indent=2),
        encoding="utf-8")
    (out_dir / "glossary_update_audit.json").write_text(json.dumps(
        {"counts": result.counts(), "updated": result.updated,
         "conflicts_resolved": result.conflicts_resolved,
         "conflicts_open": result.conflicts_open,
         "unresolved": result.unresolved,
         "term_violations": result.term_violations,
         "suggestions": result.suggestions},
        ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [f"# Glossary update — from `{Path(assets_dir).name}` + PE",
             "", "| Bucket | Count |", "|---|---|"]
    lines += [f"| {name} | {count} |"
              for name, count in result.counts().items()]
    if result.conflicts_open:
        lines += ["", "## ⚠ Open conflicts (kept OLD en — decide "
                      "manually)"]
        lines += [f"- `{e['asset']}` {e['zh'][:40]!r}: "
                  + " | ".join(r[:40] for r in e["renderings"])
                  for e in result.conflicts_open[:30]]
    if result.term_violations:
        lines += ["", "## ⚠ Term violations (final EN lacks the decided "
                      "rendering)"]
        lines += [f"- `{e['asset']}` [{e['origin']}] {e['term']} → "
                  f"{e['expected_en']!r} missing in: {e['en']}"
                  for e in result.term_violations[:40]]
    if result.suggestions:
        lines += ["", "## Suggested new term rulings (mined, not yet "
                      "decided by anyone)"]
        lines += [f"- {s['old_en']!r} → {s['new_en']!r} "
                  f"(×{s['occurrences']})" for s in result.suggestions]
    md = out_dir / "glossary_update_audit.md"
    md.write_text("\n".join(lines), encoding="utf-8")
    write_review_xlsx(result, out_dir / "glossary_review.xlsx")
    return md

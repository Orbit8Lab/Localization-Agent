"""Chat orchestrator — the natural-language interface layer OVER the
deterministic Job Controller.

Two-layer contract (design §1/§7 preserved exactly):

- The orchestrator LLM interprets the operator's text and decides WHICH
  controller tools to call. It cannot skip a stage, write an artifact, or
  clear a gate on its own initiative, because those capabilities simply do
  not exist in its tool set — a missing tool is a guarantee, a prompt rule
  is only a suggestion. `approve` exists but records the OPERATOR (the
  human named at chat start) and the controller still validates gate order.
- Execution stays deterministic: every tool is a thin wrapper over the
  same `Job` API the CLI uses.

The loop is provider-agnostic (works on any `Provider.complete`): each
step the model returns ONE JSON `ToolCall`; observations are appended to
the transcript; `respond` ends the turn. Hard cap on steps per turn.
"""
from __future__ import annotations

import json
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Dict, List, Optional, Sequence

from pydantic import Field

from .controller import GATE_NAMES, Job
from .llm import Provider, complete_json
from .context import (Block, ContextAssembler, REQUEST_LABEL, TIER_EPISODIC,
                      TIER_EVIDENCE, TIER_PLAYBOOK, TIER_SYSTEM, TIER_TASK,
                      estimate_tokens, history_blocks)
from .episodic import EpisodicMemory
from .memory import RunDB
from .schemas import IntakeBrief, Strict
from .skill_docs import SkillLibrary
from .tenancy import TenantError, resolve_read, resolve_write

# Tool calls one turn may make. 20 covers a legitimate multi-file session
# — list a drop, inspect five spreadsheets, standardize four locales,
# verify each — with headroom. Deliberately not much higher: both loops
# seen in practice made ZERO progress, so a bigger budget would have
# bought more wasted calls and a longer wait, not a better answer. The
# repeat-call breaker below is what actually stops those.
MAX_STEPS_PER_TURN = 20

# How many times one tool may return an IDENTICAL result (same tool, same
# args, same output) before the turn stops.
#
# Two was too tight for SUCCESS: re-reading a directory listing before
# acting on it is ordinary planning, and cutting the turn there stopped
# real work. Three is the point where it is a loop rather than a
# re-read. Failures keep the tighter bound — an identical error cannot
# become a different one, so a second attempt is already pointless.
REPEAT_FAILURE_LIMIT = 2
REPEAT_SUCCESS_LIMIT = 3

# The context budget. Owned in ONE place (context.ContextAssembler) rather
# than emerging from constants that never knew about each other: before
# this, SYSTEM (~2.7k) + OBSERVATION_LIMIT × MAX_STEPS_PER_TURN could reach
# ~50k tokens with history on top, and nothing in the code could say so.
#
# 100k is a deliberately CONSERVATIVE flat value: every current model
# serves at least this, so the budget can never promise room the API
# refuses. It is not this pipeline's default model's actual window — that
# number is not recorded anywhere in the codebase, and guessing it here
# would put a second unmeasured constant into a system whose whole
# complaint is unmeasured constants. Raise it per-model (or make it
# provider-derived) once the real window is known; `over_budget` on the
# assembled result is the signal if it is ever set too high.
CONTEXT_BUDGET_TOKENS = 100_000
REPLY_RESERVE_TOKENS = 2_000      # headroom for the model's own answer


def max_steps_per_turn() -> int:
    """Step cap, overridable via ``$ORBIT8_MAX_STEPS``.

    Tunable per box for the same reason as the context budget: what counts
    as "a long turn" depends on how many locales and files a project has,
    and that should not require editing source.
    """
    import os
    raw = os.environ.get("ORBIT8_MAX_STEPS")
    if not raw:
        return MAX_STEPS_PER_TURN
    try:
        value = int(raw)
    except ValueError:
        return MAX_STEPS_PER_TURN
    return value if value > 0 else MAX_STEPS_PER_TURN


def context_budget() -> int:
    """The token budget, overridable via ``$ORBIT8_CONTEXT_BUDGET``.

    An env override exists because this number is provisional: it should
    track the model's real window, and changing that must not require
    editing source. An unparseable or non-positive value falls back to the
    conservative default rather than raising — a bad env var should not
    make the chat interface unusable.
    """
    import os
    raw = os.environ.get("ORBIT8_CONTEXT_BUDGET")
    if not raw:
        return CONTEXT_BUDGET_TOKENS
    try:
        value = int(raw)
    except ValueError:
        return CONTEXT_BUDGET_TOKENS
    return value if value > 0 else CONTEXT_BUDGET_TOKENS

OBSERVATION_LIMIT = 12000         # chars of tool output fed back to the model
# inspect_file's window must stay UNDER OBSERVATION_LIMIT: the outer cap
# applies to the whole JSON envelope (filename, counts, offsets), so a
# window sized at the cap would have its tail clipped again a layer later
# and the paging offsets would lie.
INSPECT_TEXT_LIMIT = 8000         # default inspect_file text window
INSPECT_TEXT_MAX = 10000          # ceiling a caller may request


class ToolCall(Strict):
    tool: str
    args: Dict[str, object] = Field(default_factory=dict)
    message: Optional[str] = None      # required when tool == "respond"


SYSTEM = """You are Orbit8, the operator interface of a game-localization \
job. You interpret the operator's request, call tools to inspect or advance \
the job, then answer. The pipeline itself is deterministic — you only \
decide which tool to call next.

TOOLS (respond with ONE JSON object per step, no prose, no fences):
- {"tool": "status", "args": {}} — job phase, gate states, per-locale counts
- {"tool": "next_step", "args": {}} — run the job's next stage step; stops \
at gates automatically
- {"tool": "approve", "args": {"gate": "G0".."G5"}} — record the operator's \
gate approval
- {"tool": "read_artifact", "args": {"stage": 0-7, "name": "<artifact>"}} — \
read one artifact envelope
- {"tool": "list_artifacts", "args": {"stage": 0-7}} — list a stage's files
- {"tool": "flagged", "args": {"locale": "<loc>", "limit": 10}} — strings \
awaiting human review (MTPE/flagged)
- {"tool": "list_files", "args": {"dir": "<path>"}} — list files under a \
directory inside the project folder
- {"tool": "inspect_file", "args": {"path": "<file>", "offset": 0, \
"limit": 4000}} — read-only peek: a text window plus line/entry counts \
(use before deciding how to standardize anything). If the result says \
"truncated": true, the file continues — call again with "offset": \
<next_offset> to read on. NEVER describe content you have not actually \
read in a window; page to it first.
- {"tool": "standardize", "args": {"files": ["<file>", …], "output": \
"source_json" | "bilingual_jsonl", "out_name": "<name>", "target_lang": \
"<locale>", "columns": ["<source col>", "<target col>"], "out_dir": \
"<dir>"}} — convert received files into a pipeline format. source_json: \
flat {key: source text} translation input. bilingual_jsonl: LQA/MT input \
of source+target pairs. ALWAYS inspect_file first; the layout decides the \
call:
  · one file holding BOTH languages in columns → pass that one file;
  · SEPARATE files per language (e.g. "Game (Source).xlsx" + \
"Game_ja.xlsx") → pass BOTH, source first — one file alone cannot be \
paired with itself;
  · one sheet with SEVERAL target languages (a term list: English / 简体\
中文 / 日本語 …) → add "columns": ["English", "简体中文"] to pick the two \
to pair, and repeat per locale.
  "target_lang" is REQUIRED when the job has more than one target locale. \
"out_dir" writes elsewhere than the job's exports/ (e.g. \
"40-reference/glossary" for terminology); it must stay inside the project.
- {"tool": "analyze", "args": {"files": ["<file>", …], "classify": \
false}} — corpus text analysis: total/unique strings, word counts (CJK \
chars count as words), placeholders, and the domain breakdown with \
story-lines (dialogue+marketing) vs instructions (ui+system) rollup. \
Classification is reused from this job's earlier runs at no cost; set \
"classify": true only if strings are still unlabeled.
- {"tool": "compare_po", "args": {"old": "<previous .po>", "new": \
"<newly received .po>", "out_dir": "<optional report dir>"}} — diff a new \
.po drop against the previous version: added/removed keys, changed \
sources, modified translations, plus red flags (translations LOST — \
filled before, empty now — and STALE — source changed but translation \
did not). ALWAYS run this when the operator mentions receiving an \
updated/new .po for a file we already have; surface the red flags and \
word counts (the incremental work basis) in your answer. The result \
lists the actual changed entries under "source_changed" — quote those \
keys when asked WHICH entry changed; never infer a key from a count.
- {"tool": "update_glossary", "args": {"assets": "<dir with zh/en_asset\
.json + dedup_index.json>", "pe_po": "<post-edited bilingual .po>", \
"decisions": "<optional decisions .xlsx (映射留档/翻译裁定)>", "terms": \
{"<source term>": "<target rendering>", …}, "out_dir": "<output dir>", \
"distill": true}} — \
refresh the \
asset-pair glossary from post-editing results. Deterministic: assets \
join to PE strings via the dedup index; conflicting duplicate renderings \
resolve by decision compliance then majority, else stay flagged; a term \
audit lists every entry violating a decided rendering; frequent \
undocumented old→new EN replacements are mined as suggestions. With \
"distill": true, additionally boil everything down to a compact \
term-level glossary (glossary_terms.xlsx/.json, pipeline T1 shape) — \
use when the operator wants a clean glossary list rather than \
item-by-item review. Surface \
open conflicts, violations and suggestions to the operator — they are \
review work, not errors to hide.
- {"tool": "translate_po", "args": {"po": "<received bilingual .po>", \
"glossary": "<optional — omit to use the project's active termbase at \
40-reference/glossary/glossary_terms.json>", "out_dir": "<work folder>", \
"game": "<name>", "batch_size": 12, "reuse_from": "<previous translated \
.po, optional>"}} — translate ONLY the untranslated strings of a \
received drop, glossary-constrained (locked terms + family rules are \
law; exact-hit strings prefill with zero LLM cost; violators get one \
repair retry, survivors stay flagged). ALWAYS pass "reuse_from" when a \
previous translation run or delivery of the same file exists (list_files \
the work folders to find it): matching entries — same key, then \
identical source text — are carried over verbatim instead of \
re-translated, so only genuinely NEW strings cost LLM calls. When the \
operator says 已翻译的部分直接提取/复用 they mean exactly this. \
DeepSeek. Outputs a stream-patched copy (untouched entries \
byte-identical), the standard MTPE form for post-editing, and a report. \
This is a WORK PRODUCT, not a delivery — deliveries go through \
deliver_po after post-editing.
- {"tool": "scan_po", "args": {"po": "<translated bilingual .po>", \
"glossary": "<optional — defaults to the project's active termbase>", \
"out_dir": "<work folder>", "game": "<name>", \
"deterministic_only": false}} — run a FULL LQA scan on any translated \
.po without a job pipeline: the same tier cascade (T1 mechanical → T2 \
consistency → T3 LLM semantic → verify) with tier stamps and a verified \
cascade ledger, plus a whole-file check for the same source shipped with \
DIFFERENT renderings. Outputs a client bug report xlsx, an LQA PE form \
for the post-editors, a tech summary and a scan report. Use this \
whenever the operator asks to LQA / QA / 扫描 a .po file — do NOT try to \
standardize it or run job stages first. "deterministic_only": true \
skips all LLM calls (T1+T2 only).
- {"tool": "add_glossary_terms", "args": {"glossary": \
"<glossary_terms.json>", "terms": {"<source term>": "<rendering>", …}, \
"origin": "operator <date>", "force": false}} — the operator dictating \
terms directly ("把X加进术语表", "add these terms"). Each becomes a \
LOCKED entry with provenance; an unlocked mined entry is overwritten; \
an existing RULING is NOT overwritten — it comes back as a conflict for \
the operator to confirm (then re-call with "force": true). Alias forms \
share one key separated by "/"; a rename uses "old>new" as the key. A \
.bak is kept and the xlsx view is re-rendered. ALWAYS report conflicts \
verbatim and ask before forcing.
- {"tool": "extract_glossary", "args": {"po": ["<bilingual .po>", …], \
"decisions": "<optional decisions .xlsx>", "terms": \
{"<source term>": "<rendering>", …}, "out_dir": "<output dir>", \
"min_freq": 3, \
"llm_filter": false, "game": "<name>"}} — corpus-first term extraction \
(stages 0-3): dedup + markup stripping, Han n-gram mining over the FULL \
corpus INCLUDING sentence interiors (terms living only inside sentences \
become candidates), noise filter (LLM when llm_filter true, else \
deterministic heuristic), assembly with locked decisions. Outputs a T1 \
glossary + a TERM-level Glossary PE review (one row per term — a locked \
term violated in 7 strings is ONE row; string detail stays in the \
audit) + a Needs-EN work queue. Prefer this over update_glossary when \
the operator wants a glossary BUILT from a corpus or complains the \
review has sentence-length entries.
- {"tool": "deliver_po", "args": {"review": "<filled review .xlsx>", \
"po_files": ["<shipped .po>", …], "out_dir": "<optional, default \
30-deliverables>", "timestamp": "<optional YYYYMMDD>"}} — apply the \
post-editing team's decisions to the .po files and write a timestamped \
delivery folder with a delivery report. The review workbook may be \
EITHER a filled standard PE form (MTPE / LQA PE: StringID + PE_Decision \
+ PE_Modification — the shape translate_po and the LQA stage emit) OR a \
reviewed client bug report (Location/String ID + Decision + Modify \
Version); the shape is detected automatically, so never ask the \
operator to add or rename columns. Accept applies the proposed target, \
Reject&Modification applies the reviewer's text, Keep-as-it-is / \
Cannot Answer / blank leave the string alone. Deterministic; \
accepted pairs write back to the job TM as human-confirmed. Undecided/\
conflicting rows are never applied — report them to the operator. The \
po_sanity gate runs on every output by default (format/import safety, \
header label + branding, summary vs source); if the result says \
"blocked": true the delivery MUST NOT be sent — tell the operator why.
- {"tool": "respond", "args": {}, "message": "<your answer>"} — end the \
turn and reply to the operator

RULES:
- The project's ACTIVE glossary is \
40-reference/glossary/glossary_terms.json. Prefer it over any \
glossary_terms.json inside a dated 20-work/ run folder — those are the \
audit record of the round that produced a glossary, not the current \
termbase. Omit the "glossary" argument to let the tool resolve it; if a \
result carries a "glossary_notes" WARNING, relay it and suggest \
`orbit8 glossary promote`.
- ALWAYS end the turn with "respond".
- Call "approve" ONLY when the operator explicitly asked to approve that \
gate in their CURRENT message. Never approve to unblock yourself; when a \
gate is pending, describe what needs review and stop.
- "next_step" runs ONE step. Chain several when the operator asked to \
advance the job, but stop the moment a gate appears or a step fails.
- Answer in the operator's language. Be concrete: quote phases, counts, \
and finding messages from tool output. Never invent tool results.
- If a tool errors, report the error honestly and stop."""


class ChatOrchestrator:
    def __init__(self, job: Job, provider: Provider, *, operator: str,
                 provider_factory=None, dry_run: bool = False,
                 on_action: Optional[Callable[[str, str], None]] = None,
                 on_start: Optional[Callable[[str, dict], None]] = None,
                 trace_path: Optional[Path] = None,
                 skills: Optional["SkillLibrary"] = None,
                 assembler: Optional[ContextAssembler] = None,
                 episodic: Optional[EpisodicMemory] = None):
        self.job = job
        self.provider = provider
        self.operator = operator
        # PLAN §8: stage playbooks, selected by what the Controller derives
        # rather than by the operator's phrasing. Optional — a session with
        # no library behaves exactly as before.
        self.skills = skills
        self.provider_factory = provider_factory
        self.dry_run = dry_run
        self.on_action = on_action or (lambda tool, result: None)
        # fired BEFORE a tool runs, so a long call is visibly running
        # rather than indistinguishable from a hang
        self.on_start = on_start or (lambda tool, args: None)
        self.history: List[tuple] = []      # (role, text)
        # The Context Manager: ONE owner for what the model sees, replacing
        # three constants that each capped a different thing and none of
        # which knew the others existed. The system prompt is charged
        # against the budget here because the provider sends it separately.
        self.assembler = assembler or ContextAssembler(
            budget_tokens=context_budget(),
            reserve_tokens=REPLY_RESERVE_TOKENS + estimate_tokens(SYSTEM))
        self.episodic = episodic
        # debugging surface: full tool args + untruncated results, in
        # memory for /debug and on disk for post-mortem (JSONL).
        self.trace: List[dict] = []
        self.trace_path = Path(trace_path) if trace_path else None
        self.turn_no = 0

    # ------------------------------------------------------------- tools

    def _t_status(self, args: dict) -> str:
        control = self.job.control
        stage = self.job.derive()
        intake = self.job.store.read(0, "intake", IntakeBrief)
        counts = {}
        for locale in intake.target_locales:
            path = self.job.store.run_db_path(locale)
            if path.exists():
                counts[locale] = RunDB(path).counts()
        return json.dumps({
            "phase": stage.phase, "action": stage.action,
            "pending_gate": stage.gate,
            "gate_name": GATE_NAMES.get(stage.gate) if stage.gate else None,
            "target": stage.target,
            "approvals": control["approvals"], "counts": counts},
            ensure_ascii=False, default=str)

    def _t_next_step(self, args: dict) -> str:
        stage = self.job.next_step(self.provider_factory,
                                   dry_run=self.dry_run)
        after = self.job.derive()
        return json.dumps({
            "ran": None if stage.gate else
                   {"phase": stage.phase, "action": stage.action,
                    "target": stage.target},
            "now": {"phase": after.phase, "action": after.action,
                    "pending_gate": after.gate}}, ensure_ascii=False)

    def _t_approve(self, args: dict) -> str:
        gate = str(args.get("gate", ""))
        self.job.approve(gate, by=self.operator,
                         note="approved via orbit8 chat")
        return f"{gate} approved by {self.operator}"

    def _t_read_artifact(self, args: dict) -> str:
        stage = int(args.get("stage", -1))
        name = str(args.get("name", ""))
        path = self.job.store.stage_dir(stage) / f"{name}.json"
        if not path.exists():
            return f"error: no artifact {name!r} in stage {stage}"
        return path.read_text(encoding="utf-8")

    def _t_list_artifacts(self, args: dict) -> str:
        stage = int(args.get("stage", -1))
        directory = self.job.store.stage_dir(stage)
        if not directory.exists():
            return "[]"
        return json.dumps(sorted(p.name for p in directory.iterdir()))

    def _t_flagged(self, args: dict) -> str:
        locale = str(args.get("locale", ""))
        limit = int(args.get("limit", 10))
        path = self.job.store.run_db_path(locale)
        if not path.exists():
            return f"error: no run DB for locale {locale!r}"
        rows = RunDB(path).by_status("flagged", "mtpe")[:limit]
        return json.dumps([{
            "uid": r["uid"], "source": r["text"], "target": r["target"],
            "domain": r["domain"], "reason": r["resolution"],
            "findings": [f.message for f in r["findings"]]}
            for r in rows], ensure_ascii=False)

    # ------------------------------------------------------- file tools

    @property
    def project_root(self) -> Path:
        """The folder holding this job's jobs root."""
        return self.job.store.root.resolve().parent

    @property
    def tenant_id(self) -> str:
        """The organization this job belongs to — the confidentiality
        boundary for file access (see tenancy.py)."""
        try:
            return self.job.control.get("tenant_id") or "default"
        except Exception:                      # pragma: no cover - guard
            return "default"

    def _confine(self, raw: str) -> Path:
        """Resolve a path for WRITING (and for reads that should not
        cross a project boundary).

        Stays inside this job's project folder, always. Cross-project
        reads are a deliberate feature (`_confine_read`); cross-project
        WRITES are not — even within one organization, a job writing into
        another project's tree modifies assets nobody asked it to touch,
        and 30-deliverables/ is meant to be immutable.
        """
        return resolve_write(raw, project_root=self.project_root)

    def _confine_read(self, raw: str) -> Path:
        """Resolve a path for READING, allowing same-organization
        siblings.

        An agent that cannot see the org's other projects cannot learn
        from their glossaries, style guides or prior decisions — every
        project starts from nothing. But the boundary is the ORGANIZATION
        (`tenant_id`), not the folder: a sibling belonging to another
        client is refused, and so is one whose owner cannot be confirmed.
        Unmarked is treated as foreign, never as public.
        """
        return resolve_read(raw, project_root=self.project_root,
                            tenant_id=self.tenant_id)

    def _t_list_files(self, args: dict) -> str:
        directory = self._confine_read(str(args.get("dir", "")))
        if not directory.is_dir():
            return f"error: not a directory: {directory}"
        entries = [(p.name + "/" if p.is_dir() else
                    f"{p.name} ({p.stat().st_size}B)")
                   for p in sorted(directory.iterdir())
                   if p.name not in (".DS_Store",)]
        return json.dumps(entries[:60], ensure_ascii=False)

    @staticmethod
    def _tabular_preview(path: Path, suffix: str) -> Dict[str, object]:
        """Sheets, headers and a few rows — what a spreadsheet actually is.

        The question an operator asks of a localization spreadsheet is
        always "which column is the source and which is the target", so
        that is what this answers. Column letters are included because the
        adapter-writer prompt refers to them.
        """
        preview: Dict[str, object] = {"format": suffix.lstrip(".")}
        try:
            if suffix in (".csv", ".tsv"):
                import csv as _csv
                delimiter = "\t" if suffix == ".tsv" else ","
                with path.open("r", encoding="utf-8-sig",
                               errors="replace", newline="") as handle:
                    rows = []
                    for index, row in enumerate(_csv.reader(
                            handle, delimiter=delimiter)):
                        if index > 5:
                            break
                        rows.append([cell[:60] for cell in row])
                preview.update(header=rows[0] if rows else [],
                               sample_rows=rows[1:])
                return preview

            from openpyxl import load_workbook
            # read_only + values_only: a localization export can be tens of
            # MB, and nothing here needs formatting or formulas.
            book = load_workbook(path, read_only=True, data_only=True)
            preview["sheets"] = book.sheetnames
            sheets = []
            for name in book.sheetnames[:3]:
                sheet = book[name]
                rows = []
                for index, row in enumerate(
                        sheet.iter_rows(max_row=6, values_only=True)):
                    rows.append(["" if cell is None else str(cell)[:60]
                                 for cell in row])
                    if index >= 5:
                        break
                sheets.append({
                    "sheet": name,
                    "rows_total": sheet.max_row,
                    "columns": [chr(65 + n) if n < 26 else f"col{n}"
                                for n in range(len(rows[0]))] if rows else [],
                    "header": rows[0] if rows else [],
                    "sample_rows": rows[1:],
                })
            book.close()
            preview["sheet_preview"] = sheets
        except Exception as err:
            # Report the failure rather than falling back to a byte peek:
            # binary noise is what caused the loop this method exists to
            # prevent.
            preview["error"] = f"could not read as a spreadsheet: {err}"
        return preview

    def _t_inspect_file(self, args: dict) -> str:
        path = self._confine_read(str(args.get("path", "")))
        raw = path.read_bytes()
        info: Dict[str, object] = {"file": path.name, "bytes": len(raw)}
        suffix = path.suffix.lower()
        if suffix == ".po":
            from .exports import read_po_entries
            entries = read_po_entries(path)
            filled = sum(1 for _, _, t, _ in entries if t.strip())
            info.update(entries=len(entries), msgstr_filled=filled,
                        sample=[{"key": k[:16], "msgid": s[:60],
                                 "msgstr": t[:60]}
                                for k, s, t, _ in entries[:3]])
        elif suffix in (".xlsx", ".xlsm", ".csv", ".tsv"):
            # A spreadsheet is not text. Falling through to the byte peek
            # handed the model `PK\x03\x04…` (an xlsx is a zip), which
            # answers nothing about the columns — so it called inspect
            # again, and again, and burned the whole step budget on a tool
            # that "succeeded" every time.
            info.update(self._tabular_preview(path, suffix))
        else:
            # Text peek. The old flat 800-byte cap silently hid the tail of
            # every generated report (a compare .md is ~5KB, its .json
            # ~30KB), so a model asked "which entry?" saw only the summary
            # header and had to guess. ``offset``/``limit`` let it page to
            # the answer instead, and ``truncated`` says so out loud —
            # unreported truncation is what turns a partial read into a
            # confident wrong claim.
            text = raw.decode("utf-8-sig", "replace")
            offset = max(0, int(args.get("offset") or 0))
            limit = int(args.get("limit") or INSPECT_TEXT_LIMIT)
            limit = max(1, min(limit, INSPECT_TEXT_MAX))
            chunk = text[offset:offset + limit]
            info.update(chars=len(text), offset=offset, head=chunk)
            end = offset + len(chunk)
            if end < len(text):
                info["truncated"] = True
                info["next_offset"] = end
                info["remaining_chars"] = len(text) - end
        return json.dumps(info, ensure_ascii=False)

    @staticmethod
    def _read_table(path: Path):
        """(header, rows) from a spreadsheet, or None if not one.

        Deterministic, no model call. A term sheet is a plain table with
        named columns — using a generated adapter for it would key the
        cache by file suffix while the script hardcodes one language's
        column, so the ja adapter would be reused for ko and quietly emit
        Japanese.
        """
        suffix = path.suffix.lower()
        try:
            if suffix in (".csv", ".tsv"):
                import csv as _csv
                with path.open("r", encoding="utf-8-sig", errors="replace",
                               newline="") as handle:
                    rows = list(_csv.reader(
                        handle, delimiter="\t" if suffix == ".tsv" else ","))
                return (rows[0], rows[1:]) if rows else None
            if suffix in (".xlsx", ".xlsm"):
                from openpyxl import load_workbook
                book = load_workbook(path, read_only=True, data_only=True)
                sheet = book[book.sheetnames[0]]
                rows = [["" if cell is None else str(cell) for cell in row]
                        for row in sheet.iter_rows(values_only=True)]
                book.close()
                return (rows[0], rows[1:]) if rows else None
        except Exception:
            return None
        return None

    def _standardize_all_locales(self, files, intake, out_dir: Path,
                                 source_column: str, column_map: dict,
                                 stem: str) -> str:
        """Every locale from one multi-language sheet, in one call.

        Output stays one file per locale because that is what the LQA
        cascade consumes: T1 checks placeholders and width against ONE
        target, T2 checks consistency within ONE locale, T3 reviews ONE
        language pair. A merged file could not say which locale a finding
        belongs to.

        What this removes is the WASTE, not the format: reading the same
        sheet four times, and re-deriving how to read it four times.

        Failures are per-locale — one bad column must not cost the others,
        and the operator needs to know exactly which one to fix.
        """
        from .exports import emit_bilingual_jsonl

        # Read the sheet ONCE with the deterministic reader. A term sheet
        # is a plain table with named columns, so there is nothing here
        # that needs a generated adapter — and using one would key the
        # cache by suffix while the script itself hardcodes a column,
        # which is how the wrong language ends up in a reused adapter.
        table = self._read_table(files[0]) if len(files) == 1 else None

        results, failures = [], []
        if table is not None:
            header, rows = table
            if source_column not in header:
                return json.dumps({
                    "status": "failed", "written": [],
                    "failed": [f"source column {source_column!r} not found; "
                               f"headers are {header}"]}, ensure_ascii=False)
            source_index = header.index(source_column)
            for locale, column in column_map.items():
                if locale not in intake.target_locales:
                    failures.append(
                        f"{locale}: not a target locale of this job "
                        f"({', '.join(intake.target_locales)})")
                    continue
                if str(column) not in header:
                    failures.append(
                        f"{locale}: column {column!r} not found; headers "
                        f"are {header}")
                    continue
                target_index = header.index(str(column))
                pairs = [(row[source_index], row[source_index],
                          row[target_index], str(files[0]))
                         for row in rows
                         if len(row) > max(source_index, target_index)
                         and str(row[source_index]).strip()]
                out = out_dir / f"{stem}_{intake.source_lang}-{locale}.jsonl"
                try:
                    from .exports import _write_pairs
                    empty = sum(1 for _k, _s, t, _l in pairs
                                if not str(t).strip())
                    written, _ = _write_pairs(
                        pairs, out, source_lang=intake.source_lang,
                        target_lang=locale, empty=empty, identical=0)
                except Exception as err:
                    failures.append(f"{locale} (column {column!r}): {err}")
                    continue
                results.append({"locale": locale, "column": column,
                                "written": written,
                                "empty_targets_included": empty,
                                "path": str(out)})
            return json.dumps({
                "status": "complete" if results and not failures
                          else ("partial" if results else "failed"),
                "written": results, "failed": failures,
                "next": ("Each locale has its own file — the LQA cascade "
                         "audits ONE language pair at a time. Do not "
                         "re-run these; report the paths and counts."
                         if results else
                         "Nothing was written; fix the errors above.")},
                ensure_ascii=False)
        for locale, column in column_map.items():
            if locale not in intake.target_locales:
                failures.append(
                    f"{locale}: not a target locale of this job "
                    f"({', '.join(intake.target_locales)})")
                continue
            out = out_dir / f"{stem}_{intake.source_lang}-{locale}.jsonl"
            try:
                written, empty = emit_bilingual_jsonl(
                    files, out, source_lang=intake.source_lang,
                    target_lang=locale,
                    columns=[source_column, str(column)],
                    fallback=self.job._bilingual_fallback(self.provider,
                                                          self.dry_run))
            except Exception as err:
                failures.append(f"{locale} (column {column!r}): {err}")
                continue
            results.append({"locale": locale, "column": column,
                            "written": written,
                            "empty_targets_included": empty,
                            "path": str(out)})
        return json.dumps({
            "status": "complete" if results and not failures
                      else ("partial" if results else "failed"),
            "written": results, "failed": failures,
            "next": ("Each locale has its own file — the LQA cascade "
                     "audits ONE language pair at a time. Do not re-run "
                     "these; report the paths and counts."
                     if results else
                     "Nothing was written; fix the errors above.")},
            ensure_ascii=False)

    def _t_standardize(self, args: dict) -> str:
        from .exports import emit_bilingual_jsonl, emit_flat_json
        from .ingest import ingest_any
        from .schemas import IntakeBrief
        files = [self._confine(str(f)) for f in args.get("files", [])]
        output = str(args.get("output", ""))
        if not files:
            return "error: no files given"
        intake = self.job.store.read(0, "intake", IntakeBrief)
        # Where the output lands is a real decision, not a constant. A
        # glossary belongs in the PROJECT's 40-reference/glossary/, where
        # every later stage resolves it (project_paths.py); a job's
        # exports/ is the right home only for job-scoped conversions.
        # Hardcoding it forced the agent to say "the tool will not let me
        # put this where you asked" — accurate, and a missing capability.
        # `_confine` (write path) keeps it inside this project.
        if args.get("out_dir"):
            out_dir = self._confine(str(args["out_dir"]))
        else:
            out_dir = self.job.store.job_dir / "exports"
        if output == "source_json":
            name = str(args.get("out_name") or "strings")
            if not name.endswith(".json"):
                name += ".json"
            out = out_dir / name
            fallback = self.job._adapter_fallback(self.provider,
                                                  self.dry_run)
            records = []
            for file in files:
                records.extend(ingest_any(file, fallback=fallback))
            count = emit_flat_json(records, out)
            return json.dumps({
                "status": "complete", "written": count, "path": str(out),
                "next": ("This file is DONE — do not call standardize for "
                         "it again. Report the path and count to the "
                         "operator.")})
        if output == "bilingual_jsonl":
            # `column_map` supplies the locales itself (one per column), so
            # it is handled BEFORE the single-locale rules below — asking
            # it for one `target_lang` would be asking the wrong question.
            column_map = args.get("column_map")
            if column_map:
                if not isinstance(column_map, dict):
                    return ('error: "column_map" maps locale → column '
                            'name, e.g. {"ja": "日本語", "ko": "한국어"}')
                source_column = str(args.get("source_column") or "")
                if not source_column:
                    return ('error: "column_map" needs "source_column" too '
                            '— which column holds the source language')
                return self._standardize_all_locales(
                    files, intake, out_dir, source_column, column_map,
                    str(args.get("out_name") or "glossary"))

            # Never guess the locale when several are configured. Falling
            # back to target_locales[0] labelled a Japanese file "zh-CN"
            # and wrote 400 rows nobody could use — a mislabelled export
            # is worse than a refusal, because it looks fine.
            locale = str(args.get("target_lang") or "")
            if not locale:
                if len(intake.target_locales) != 1:
                    return (f"error: this job has "
                            f"{len(intake.target_locales)} target locales "
                            f"({', '.join(intake.target_locales)}); pass "
                            f'"target_lang" to say which one this file '
                            f"holds")
                locale = intake.target_locales[0]
            if locale not in intake.target_locales:
                return (f"error: {locale!r} is not a target locale of this "
                        f"job ({', '.join(intake.target_locales)})")
            name = str(args.get("out_name")
                       or f"pairs_{intake.source_lang}-{locale}")
            if not name.endswith(".jsonl"):
                name += ".jsonl"
            out = out_dir / name
            # Cheap structural check BEFORE any model call. A single
            # non-.po file cannot yield pairs unless it holds both
            # languages in columns — and when it does not, the adapter
            # writer spends three attempts discovering that a file cannot
            # be paired with itself (~190s observed). A .po is exempt: it
            # carries msgid AND msgstr, so one file is genuinely enough.
            # `columns` names which two columns to pair when one sheet
            # holds several languages — a term sheet
            # (English/简体中文/繁體中文/한국어/日本語) is one row per
            # concept with FOUR target columns, so "source + target file"
            # does not describe it and neither pipeline format fits. With
            # the columns named, one sheet yields one pair set per locale.
            columns = args.get("columns") or []
            single = [f for f in files if f.suffix.lower() != ".po"]

            # A multi-language sheet maps to SEVERAL locales at once.
            # The OUTPUT must stay one source+target per file — the LQA
            # cascade checks placeholders, consistency and semantics
            # against ONE target, and a merged report could not say which
            # locale a finding belongs to. But re-reading the same sheet
            # four times to say that is pure waste, so one call can emit
            # every locale.
            if columns and len(columns) != 2:
                return ('error: "columns" takes exactly two names — the '
                        'source column and the target column, e.g. '
                        '["English", "简体中文"]')
            if len(files) == 1 and single and not columns:
                preview = self._tabular_preview(files[0],
                                                files[0].suffix.lower())
                header = (preview.get("header")
                          or (preview.get("sheet_preview") or [{}])[0]
                          .get("header") or [])
                if len(header) < 3:
                    return (
                        f"error: {files[0].name!r} has {len(header)} "
                        f"column(s) — too few to hold both a source and a "
                        f"{locale} translation. If the source text is in a "
                        f"separate file, pass BOTH (source first, then the "
                        f"{locale} file) so they can be paired on key.")
            written, empty = emit_bilingual_jsonl(
                files, out, source_lang=intake.source_lang,
                target_lang=locale, columns=list(columns) or None,
                fallback=self.job._bilingual_fallback(self.provider,
                                                      self.dry_run))
            # `status: complete` and the explicit instruction are not
            # decoration: a bare {"written": 1395, "path": …} reads as
            # ambiguous, and the model re-issued the identical call three
            # times before the loop breaker stopped it — repeating work
            # that had already succeeded on the first try.
            return json.dumps({
                "status": "complete",
                "written": written,
                "empty_targets_included": empty,
                "path": str(out),
                "next": (f"This file is DONE — do not call standardize for "
                         f"{locale} again. Report the path and counts to "
                         f"the operator, or move on to another locale.")})
        return (f"error: unknown output {output!r}; use source_json or "
                f"bilingual_jsonl")

    def _t_update_glossary(self, args: dict) -> str:
        from .glossary_update import (TermDecision, load_decisions_xlsx,
                                      refresh_glossary,
                                      write_update_outputs)
        assets = self._confine(str(args.get("assets", "")))
        pe_po = self._confine(str(args.get("pe_po", "")))
        out_dir = self._confine(str(args.get("out_dir", "")))
        decisions = (load_decisions_xlsx(
            self._confine(str(args["decisions"])))
            if args.get("decisions") else [])
        for zh, en in dict(args.get("terms") or {}).items():
            decisions.append(TermDecision(zh=str(zh), en=str(en)))
        result = refresh_glossary(assets, pe_po, decisions)
        md = write_update_outputs(result, out_dir, assets)
        distilled = {}
        if args.get("distill"):
            from .glossary_update import (distill_term_glossary,
                                          write_term_glossary)
            zh_asset = json.loads(
                (Path(assets) / "zh_asset.json").read_text("utf-8"))
            glossary, ties = distill_term_glossary(
                zh_asset, result.updated_en, decisions)
            path = write_term_glossary(glossary, ties, out_dir)
            distilled = {"terms_xlsx": str(path),
                         "terms_total": len(glossary["terms"]),
                         "terms_locked":
                             glossary["metadata"]["locked_terms"],
                         "ties_excluded": len(ties)}
        return json.dumps(
            {"counts": result.counts(), "audit": str(md),
             **({"distilled": distilled} if distilled else {}),
             "review_xlsx": str(Path(str(out_dir))
                                / "glossary_review.xlsx"),
             "conflicts_open": result.conflicts_open[:8],
             "term_violations": result.term_violations[:8],
             "suggestions": result.suggestions[:8]}, ensure_ascii=False)

    def _t_translate_po(self, args: dict) -> str:
        from .llm import OpenAICompatProvider, autoload_env
        from .po_translate import translate_untranslated
        autoload_env()
        provider = OpenAICompatProvider("deepseek")
        from .project_paths import resolve_glossary
        po_path = self._confine(str(args.get("po", "")))
        glossary, notes = resolve_glossary(
            hint=(self._confine(str(args["glossary"]))
                  if args.get("glossary") else None),
            start=po_path)
        if glossary is None:
            return json.dumps({"error": "no glossary resolved",
                               "notes": notes}, ensure_ascii=False)
        run = translate_untranslated(
            po_path, glossary,
            self._confine(str(args.get("out_dir", ""))),
            provider=provider, game=str(args.get("game", "")),
            locale=str(args.get("locale", "en")),
            batch_size=int(args.get("batch_size", 12)),
            reuse_from=(self._confine(str(args["reuse_from"]))
                        if args.get("reuse_from") else None))
        return json.dumps(
            {"glossary_used": str(glossary), "glossary_notes": notes,
             "entries_total": run.total, "untranslated": run.todo,
             "reused": len(run.reused),
             "prefilled": len(run.prefilled),
             "translated": len(run.translated),
             "repaired": run.repaired,
             "violations": run.violations[:10],
             "tokens_spent": run.tokens, "sanity_format": run.sanity,
             "mtpe_form": str(Path(str(args.get("out_dir", "")))
                              / "mtpe_form.xlsx")}, ensure_ascii=False)

    def _t_scan_po(self, args: dict) -> str:
        from .po_scan import scan_po
        from .project_paths import resolve_glossary
        po_path = self._confine(str(args.get("po", "")))
        glossary, notes = resolve_glossary(
            hint=(self._confine(str(args["glossary"]))
                  if args.get("glossary") else None),
            start=po_path)
        deterministic = bool(args.get("deterministic_only"))
        provider = None if (deterministic or self.dry_run) else (
            self.provider_factory(str(args.get("locale", "en")))
            if self.provider_factory else self.provider)
        result = scan_po(
            po_path, glossary,
            self._confine(str(args.get("out_dir", ""))),
            provider=provider, game=str(args.get("game", "")),
            locale=str(args.get("locale", "en")),
            source_lang=str(args.get("source_lang", "zh-CN")),
            deterministic_only=deterministic or provider is None)
        report = result.report
        return json.dumps(
            {"glossary_used": str(glossary) if glossary else None,
             "glossary_notes": notes,
             "checked": report.checked,
             "flagged_strings": report.flagged_strings,
             "findings_total": report.findings_total,
             "by_severity": report.by_severity,
             "by_bug_type": report.by_bug_type,
             "cascade_ledger": report.cascade_ledger,
             "block_ship": report.block_ship,
             "inconsistent_sources": len(result.inconsistent),
             "suggestions": len(result.suggestions),
             **result.outputs}, ensure_ascii=False)

    def _t_add_glossary_terms(self, args: dict) -> str:
        from datetime import datetime
        from .glossary_edit import TermEdit, edit_glossary_file
        edits = [TermEdit(zh=str(zh), en=str(en))
                 for zh, en in dict(args.get("terms") or {}).items()]
        if not edits:
            return "error: no terms given"
        stamp = datetime.now().strftime("%Y%m%d-%H%M")
        _g, result, backup = edit_glossary_file(
            self._confine(str(args.get("glossary", ""))), edits,
            origin=str(args.get("origin")
                       or f"operator {datetime.now():%Y-%m-%d}"),
            force=bool(args.get("force")), backup_stamp=stamp)
        return json.dumps(
            {"added": result.added, "aliased": result.aliased,
             "overwritten": result.overwritten,
             "retired": result.retired, "unchanged": result.unchanged,
             "conflicts": result.conflicts, "flagged": result.flagged,
             "backup": str(backup) if backup else None},
            ensure_ascii=False)

    def _t_extract_glossary(self, args: dict) -> str:
        from .glossary_update import TermDecision, load_decisions_xlsx
        from .term_extract import (extract_glossary,
                                   write_extraction_outputs)
        po_paths = [self._confine(str(p))
                    for p in (args.get("po") or [])]
        out_dir = self._confine(str(args.get("out_dir", "")))
        decisions = (load_decisions_xlsx(
            self._confine(str(args["decisions"])))
            if args.get("decisions") else [])
        for zh, en in dict(args.get("terms") or {}).items():
            decisions.append(TermDecision(zh=str(zh), en=str(en)))
        provider = None
        if args.get("llm_filter"):
            from .llm import OpenAICompatProvider, autoload_env
            autoload_env()
            provider = OpenAICompatProvider("deepseek")
        result = extract_glossary(
            po_paths, decisions, provider=provider,
            min_freq=int(args.get("min_freq", 3)),
            game=str(args.get("game", "")),
            locale=str(args.get("locale", "en")))
        write_extraction_outputs(result, out_dir)
        return json.dumps(
            {"stats": result.stats,
             "review_xlsx": str(Path(str(out_dir))
                                / "extract_review.xlsx"),
             "glossary_json": str(Path(str(out_dir))
                                  / "glossary_terms.json"),
             "conflicts": result.conflicts[:8],
             "violations": result.violations[:8],
             "needs_en": result.needs_en[:8]}, ensure_ascii=False)

    def _t_compare_po(self, args: dict) -> str:
        from .po_compare import compare_po, write_compare_report
        old = self._confine(str(args.get("old", "")))
        new = self._confine(str(args.get("new", "")))
        result = compare_po(old, new)
        payload = {"counts": result.counts(),
                   "needs_attention": result.needs_attention,
                   "translation_lost": result.translation_lost[:10],
                   # EVERY source edit, not only the stale ones. A count
                   # with no keys cannot answer "which entry changed?" —
                   # the model is left to hunt the report file on disk and
                   # guess, which is how a correct answer ends up
                   # indefensible.
                   "source_changed": result.source_changed[:10],
                   "stale_translations": [
                       e for e in result.source_changed
                       if e.get("stale")][:10],
                   "work_summary": result.work_summary()}
        if args.get("out_dir"):
            md = write_compare_report(
                result, self._confine(str(args["out_dir"])))
            payload["report"] = str(md)
        return json.dumps(payload, ensure_ascii=False)

    def _t_deliver_po(self, args: dict) -> str:
        from .memory import TranslationMemory
        from .po_patch import deliver_from_review
        review = self._confine(str(args.get("review", "")))
        po_files = [self._confine(str(f))
                    for f in args.get("po_files", [])]
        if not po_files:
            return "error: no po_files given"
        project_root = self.job.store.root.resolve().parent
        out_dir = (self._confine(str(args["out_dir"]))
                   if args.get("out_dir")
                   else project_root / "30-deliverables")
        intake = self.job.store.read(0, "intake", IntakeBrief)
        report = deliver_from_review(
            review, po_files, out_dir,
            timestamp=(str(args["timestamp"])
                       if args.get("timestamp") else None),
            tm=TranslationMemory(self.job.store.tm_path()),
            locale=intake.target_locales[0])
        return json.dumps(
            {"counts": report.counts(), "outputs": report.outputs,
             "report": f"{report.delivery_dir}/DELIVERY_REPORT.md",
             "blocked": report.blocked,
             "sanity": {name: {"verdict": r["verdict"],
                               "errors": r["error_details"][:5]}
                        for name, r in report.sanity.items()},
             "relabeled": report.relabeled,
             "undecided": report.undecided[:10],
             "unmatched": report.unmatched[:10],
             "conflicts": report.conflicts[:10],
             "inconsistent": report.inconsistent[:10]}, ensure_ascii=False)

    def _t_analyze(self, args: dict) -> str:
        from .analysis import analyze_corpus, labels_from_run_dbs
        from .ingest import ingest_any
        files = [self._confine(str(f)) for f in args.get("files", [])]
        if not files:
            return "error: no files given"
        fallback = self.job._adapter_fallback(self.provider, self.dry_run)
        records = []
        for file in files:
            records.extend(ingest_any(file, fallback=fallback))
        labels = labels_from_run_dbs(self.job.store.job_dir / "runs")
        provider = (self.provider if args.get("classify")
                    and not self.dry_run else None)
        report = analyze_corpus(records, labels=labels, provider=provider)
        return report.model_dump_json()

    # -------------------------------------------------------------- loop

    def _tools(self) -> Dict[str, Callable[[dict], str]]:
        """The tool registry — THE list of things this agent can do.

        Hoisted out of `turn()` so it is a single source of truth: the
        skill-doc loader validates every `tools:` entry against this map
        (PLAN §8), and a doc naming a capability that does not exist fails
        at load instead of sending the agent hunting for it. A second copy
        of this list would let the two drift apart silently, which is the
        exact failure mode skill docs are supposed to stop having.
        """
        return {"status": self._t_status, "next_step": self._t_next_step,
                "approve": self._t_approve,
                "read_artifact": self._t_read_artifact,
                "list_artifacts": self._t_list_artifacts,
                "flagged": self._t_flagged,
                "list_files": self._t_list_files,
                "inspect_file": self._t_inspect_file,
                "standardize": self._t_standardize,
                "analyze": self._t_analyze,
                "compare_po": self._t_compare_po,
                "deliver_po": self._t_deliver_po,
                "update_glossary": self._t_update_glossary,
                "extract_glossary": self._t_extract_glossary,
                "add_glossary_terms": self._t_add_glossary_terms,
                "translate_po": self._t_translate_po,
                "scan_po": self._t_scan_po}

    @classmethod
    def tool_names(cls) -> frozenset:
        """Tool names without needing a live session (for the loader and
        its tests). Derived from the same method the agent uses, via a
        throwaway instance-free introspection of the handler prefix."""
        return frozenset(
            name[3:] for name in dir(cls) if name.startswith("_t_"))

    def _playbook(self) -> Optional[str]:
        """The stage playbook for the CURRENT derived stage (PLAN §8).

        Keyed on `job.derive()`, so the agent gets the playbook for the
        stage the artifact tree says it is in — it cannot request another
        stage's guidance by rephrasing. Failures are swallowed: a playbook
        is guidance layered over a working system, and a missing or broken
        doc must never take a session down.
        """
        if self.skills is None:
            return None
        try:
            stage = self.job.derive()
            skill = self.skills.for_stage(stage.phase, stage.gate)
            return skill.prompt_section() if skill else None
        except Exception:                      # pragma: no cover - guard
            return None

    def _task_block(self) -> Optional[str]:
        """Where the job actually is, from the artifact tree.

        Tier 1 — outranks even this turn's evidence, because a large tool
        result must never be able to push out the agent's knowledge of
        which stage it is in. An agent that forgets the stage calls the
        wrong stage's actions, and does it confidently.
        """
        try:
            stage = self.job.derive()
        except Exception:                      # pragma: no cover - guard
            return None
        line = f"[job state] phase={stage.phase} action={stage.action}"
        if stage.gate:
            line += (f" — WAITING on {stage.gate} "
                     f"({GATE_NAMES.get(stage.gate, '')}); only a human "
                     f"approval clears it")
        if stage.target:
            line += f" locale={stage.target}"
        return line

    def _episodic_block(self, user_msg: str) -> Optional[str]:
        """What earlier sessions of THIS job did, when the request refers
        back to something. Scoped to the job directory — never across
        jobs (see episodic.py)."""
        if self.episodic is None:
            return None
        try:
            episodes = self.episodic.recall(user_msg,
                                            exclude=self.trace_path)
            return self.episodic.as_block_text(episodes) or None
        except Exception:                      # pragma: no cover - guard
            return None

    def build_context(self, user_msg: str,
                      evidence: Optional[Sequence[str]] = None):
        """Assemble one call's context under an owned token budget.

        This replaces three independent limits that never knew about each
        other — `history[-30:]`, a per-result `OBSERVATION_LIMIT`, and the
        reply's `max_tokens`. Their sum could reach ~50k tokens with
        nothing in the code aware of it.
        """
        # SYSTEM is passed to the provider as its own argument, so it must
        # NOT be rendered into this text — it would be sent twice. It still
        # occupies the model's window, so it is charged to the budget as a
        # reservation rather than a block. A budget that ignores the system
        # prompt is short by ~2.7k tokens on every call.
        blocks: List[Block] = []

        task = self._task_block()
        if task:
            blocks.append(Block(tier=TIER_TASK, text=task, label="job-state"))

        for index, item in enumerate(evidence or ()):
            blocks.append(Block(
                tier=TIER_EVIDENCE, text=item, label=f"evidence-{index}",
                trimmable=True, order=index))

        playbook = self._playbook()
        if playbook:
            blocks.append(Block(tier=TIER_PLAYBOOK, text=playbook,
                                label="playbook"))

        recalled = self._episodic_block(user_msg)
        if recalled:
            blocks.append(Block(tier=TIER_EPISODIC, text=recalled,
                                label="episodic"))

        blocks.extend(history_blocks(self.history))
        blocks.append(Block(
            tier=TIER_EVIDENCE, order=10_000, label=REQUEST_LABEL,
            text=f"[operator] {user_msg}\nNext step — ONE JSON tool call:"))
        return self.assembler.assemble(blocks)

    def _transcript(self, user_msg: str,
                    evidence: Optional[Sequence[str]] = None) -> str:
        context = self.build_context(user_msg, evidence)
        if context.elisions:
            self._trace("context", tokens=context.tokens,
                        elisions=[e.describe() for e in context.elisions],
                        over_budget=context.over_budget)
        return context.text

    def _trace(self, event: str, **fields) -> None:
        """Append one JSONL record to the session trace. The trace is the
        debugging surface: what the model decided, with which arguments,
        and what the tool actually returned — never truncated the way the
        on_action console line is."""
        record = {"ts": datetime.now(timezone.utc).isoformat(
            timespec="seconds"), "turn": self.turn_no, "event": event,
            **fields}
        self.trace.append(record)
        if self.trace_path is None:
            return
        try:
            self.trace_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.trace_path, "a", encoding="utf-8") as handle:
                handle.write(json.dumps(record, ensure_ascii=False,
                                        default=str) + "\n")
        except OSError:
            pass                    # tracing must never break a session

    @staticmethod
    def _step_limit_reply(evidence: List[str]) -> str:
        """What was accomplished, and how to pick it up.

        The old message ("I hit the per-turn step limit... state is
        unchanged") was honest and useless: it named neither what the turn
        learned nor what to type next, so the operator restarted work the
        agent had already done. An unfinished turn should be RESUMABLE.
        """
        lines = ["I reached the per-turn step limit before finishing."]
        if evidence:
            lines.append("\nWhat I did:")
            for item in evidence:
                # Each entry reads "[tool] name({args}) -> observation".
                head = item[len("[tool] "):] if item.startswith("[tool] ") \
                    else item
                call, _, result = head.partition(" -> ")
                outcome = result.strip().replace("\n", " ")
                if outcome.startswith("error:"):
                    summary = outcome[:100]
                else:
                    summary = f"ok ({len(outcome)} chars)" if len(
                        outcome) > 120 else outcome[:120]
                lines.append(f"  {call[:90]} → {summary}")
        lines.append(
            "\nThe job state is unchanged beyond those steps. Say "
            '"continue" to resume, or narrow the request — asking for one '
            "file or one locale at a time keeps a turn well inside the "
            "limit.")
        return "\n".join(lines)

    def _retire_evidence(self, evidence: List[str]) -> None:
        """Move this turn's tool results down into history.

        Evidence is 'what happened THIS turn' — once the turn is answered
        it becomes conversation, and the next turn's tool results take the
        high tier. Without this every past turn's output would keep
        outranking the current one forever.
        """
        for item in evidence:
            self.history.append(("tool", item))
        evidence.clear()

    def turn(self, user_msg: str) -> str:
        """One operator message → tool calls → one reply."""
        self.turn_no += 1
        self._trace("operator", message=user_msg)
        tools = self._tools()
        pending_user = user_msg
        # THIS turn's tool results. They ride the EVIDENCE tier while the
        # turn is live — outranking history and trimmable under pressure —
        # and are demoted into history once the turn ends. Appending them
        # straight to history (as this loop used to) put the very thing the
        # model was called about into the lowest, non-trimmable tier.
        evidence: List[str] = []
        # (tool, args, error) -> how many times it has failed identically.
        repeats: Dict[tuple, int] = {}
        for _ in range(max_steps_per_turn()):
            self.on_start("(thinking)", {})
            think_started = time.monotonic()
            call = complete_json(self.provider, SYSTEM,
                                 self._transcript(pending_user, evidence),
                                 ToolCall,
                                 temperature=0.0, max_tokens=1200)
            self._trace("decide", tool=call.tool, args=call.args,
                        seconds=round(time.monotonic() - think_started, 2))
            if call.tool == "respond":
                reply = call.message or "(empty reply)"
                self._retire_evidence(evidence)
                self.history.append(("operator", user_msg))
                self.history.append(("orbit8", reply))
                self._trace("respond", message=reply)
                return reply
            handler = tools.get(call.tool)
            started = time.monotonic()
            failed = None
            self._trace("tool_start", tool=call.tool, args=call.args)
            self.on_start(call.tool, call.args)
            if handler is None:
                observation = (f"error: unknown tool {call.tool!r}; "
                               f"valid: {sorted(tools)} or respond")
                failed = "unknown tool"
            else:
                try:
                    observation = handler(call.args)
                except Exception as err:       # tool errors are data, not crashes
                    observation = f"error: {err}"
                    failed = f"{type(err).__name__}: {err}"
            self._trace("tool", tool=call.tool, args=call.args,
                        result=observation, error=failed,
                        seconds=round(time.monotonic() - started, 2))
            # Say so when the tail is dropped. A silently clipped
            # observation is indistinguishable from a complete one, which
            # is how a model ends up describing content it never received.
            if len(observation) > OBSERVATION_LIMIT:
                observation = (
                    observation[:OBSERVATION_LIMIT]
                    + f"\n…[truncated: {len(observation) - OBSERVATION_LIMIT}"
                      f" more chars of this result were NOT shown. Do not "
                      f"describe what is not visible above; narrow the "
                      f"request or page with offset/limit.]")
            self.on_action(call.tool, observation)
            evidence.append(
                f"[tool] {call.tool}"
                f"({json.dumps(call.args, default=str)}) -> {observation}")

            # Circuit breaker. Re-issuing a call that just failed the same
            # way cannot succeed — the inputs are identical — so repeating
            # it only burns the step budget and ends in a generic
            # "step limit" message that hides the actual error. Observed:
            # 14 identical `status` calls against a missing job, ~70s, and
            # the operator never saw the error that explained it.
            # Tools report failure two ways: by raising (caught above into
            # `failed`) and by RETURNING an "error: …" string, which most
            # of them do. Counting only exceptions would leave the second
            # kind looping exactly as before.
            # Counting only FAILURES missed the worse loop: a tool that
            # succeeds identically every time. Twelve `list_files` calls
            # with the same args each returned valid JSON, so nothing
            # tripped — and the turn died at the step limit with a generic
            # message. Same call, same result, no new information, whether
            # or not it "worked".
            outcome = failed or (observation
                                 if observation.startswith("error:")
                                 else None)
            signature = (call.tool,
                         json.dumps(call.args, sort_keys=True, default=str),
                         outcome if outcome is not None else observation)
            repeats[signature] = repeats.get(signature, 0) + 1
            limit = (REPEAT_FAILURE_LIMIT if outcome is not None
                     else REPEAT_SUCCESS_LIMIT)
            if repeats[signature] >= limit:
                verb = ("failed the same way" if outcome is not None
                        else "returned the same result")
                reply = (
                    f"Stopping: `{call.tool}` {verb} "
                    f"{repeats[signature]} times and repeating it "
                    f"cannot change that.\n\n{observation}")
                self._retire_evidence(evidence)
                self.history.append(("operator", user_msg))
                self.history.append(("orbit8", reply))
                self._trace("abort", reason="repeated_call",
                            tool=call.tool, error=outcome)
                return reply
        reply = self._step_limit_reply(evidence)
        self._retire_evidence(evidence)
        self.history.append(("operator", user_msg))
        self.history.append(("orbit8", reply))
        self._trace("abort", reason="step_limit", steps=len(evidence))
        return reply

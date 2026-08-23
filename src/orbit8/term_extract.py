"""Corpus-first glossary extraction — Stages 0–3.

Port of the staged design proven in the glossary research project
(`~/GitHub/localization/glossary/scripts/source_corpus_analysis.py`,
Stage-1 report in `results/SOURCE_ANALYSIS.md`), adapted from EN-source
to zh-source corpora:

  Stage 0  intake + normalization — per-engine markup/placeholder
           stripping, dedup FIRST (duplicates poison frequency scores)
  Stage 1  deterministic candidate mining — Han n-grams from the FULL
           corpus including sentence interiors, scored by frequency +
           standalone-string evidence + 「」《》 enclosure + EN-side
           capitalization (the bilingual pair replaces the mid-sentence-
           capitalization signal that does not exist in Chinese);
           high-recall by design
  Stage 2  noise filter — LLM pass (DeepSeek batches, keep/drop + EN
           suggestion) or a deterministic heuristic fallback when no
           provider is given
  Stage 3  assembly — locked decisions always win; standalone-rendering
           majority elects EN; per-TERM aggregation of conflicts and
           locked-term violations (瘟疫点 wrong in 7 strings = ONE review
           row with 7 pieces of evidence, not 7 rows)

Outputs: pipeline T1 glossary (docs/STANDARDS.md §2.1), a term-level
Glossary PE review form (§4.3 — short terms only, open questions only),
and an audit report. String-level facts stay in the audit, never in the
review form.
"""
from __future__ import annotations

import json
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from pydantic import BaseModel

from .exports import read_po_entries
from .glossary_update import (TermDecision, _en_base, _en_has,
                              TERM_MAX_ZH_CHARS, write_term_glossary)
from .standards import FORM_DROPDOWNS, GLOSSARY_PE_FORM_HEADERS

# ------------------------------------------------------------------ stage 0

# UE/gettext markup that poisons n-gram mining: {Var} placeholders, rich
# text/HTML-ish tags, printf tokens, entities. Text from read_po_entries is
# already unescaped, so real newlines are handled by whitespace collapse.
_MARKUP_RE = re.compile(r"\{[^{}]*\}|<[^<>]*>|%\w|&[a-z]+;")
_HAN_RUN_RE = re.compile(r"[一-鿿]{2,}")

# Function characters that never begin or end a term. Deliberately does NOT
# include productive term heads (值 点 岛 铺 …) — 瘟疫点/精神值 are terms.
_EDGE_STOP = set("的了是在有和与或之被将已从对而且并你我他她它这那此每各"
                 "另若如把让使当会能可需应仍再更最很都也还请勿即所其为个"
                 "只就要不没未非按向以及等去来到中")

_ENCLOSURE_OPEN = set("「『《【“\"'（(")
_ENCLOSURE_CLOSE = set("」』》】”\"'）)")


@dataclass
class CorpusString:
    key: str
    zh: str            # cleaned source
    en: str            # representative cleaned target ("" if untranslated)
    count: int = 1     # how many raw entries deduped into this string
    en_all: Counter = field(default_factory=Counter)  # ALL distinct EN


def strip_markup(text: str) -> str:
    """Stage-0 normalization: placeholders/tags out, whitespace collapsed.
    Per the research report, skipping this poisons term extraction."""
    return re.sub(r"\s+", " ", _MARKUP_RE.sub(" ", text)).strip()


def load_corpus(po_paths: Sequence[Path]) -> List[CorpusString]:
    """Read bilingual .po files (msgid=zh, msgstr=EN) and dedup by cleaned
    source text — dedup FIRST, so repetition can't inflate scores. Distinct
    EN renderings of the same zh are all kept (``en_all``): duplicate
    strings translated differently are the conflict signal, and dedup must
    not erase it."""
    merged: Dict[str, CorpusString] = {}
    for po_path in po_paths:
        for key, zh, en, _loc in read_po_entries(Path(po_path)):
            zh_clean = strip_markup(zh)
            if not zh_clean:
                continue
            en_clean = strip_markup(en)
            hit = merged.get(zh_clean)
            if hit is None:
                hit = merged[zh_clean] = CorpusString(
                    key=key, zh=zh_clean, en=en_clean)
            else:
                hit.count += 1
                if not hit.en:                     # any translated copy wins
                    hit.en = en_clean
            if en_clean:
                hit.en_all[en_clean] += 1
    return list(merged.values())


# ------------------------------------------------------------------ stage 1

@dataclass
class Candidate:
    zh: str
    freq: int = 0          # distinct corpus strings containing the term
    standalone: int = 0    # strings that ARE exactly the term
    enclosed: int = 0      # occurrences wrapped in 「」《》【】 quotes
    en_cap: int = 0        # standalone EN renderings that are capitalized
    renderings: Counter = field(default_factory=Counter)  # standalone EN
    examples: List[Tuple[str, str]] = field(default_factory=list)

    @property
    def score(self) -> int:
        # freq + 3×name-slot + 2×enclosure + 1×EN-capitalization — the
        # research script's weights, with zh-appropriate signals.
        return self.freq + 3 * self.standalone + 2 * self.enclosed \
            + self.en_cap


def _grams(run: str, max_n: int) -> Iterable[str]:
    for n in range(2, min(max_n, len(run)) + 1):
        for i in range(len(run) - n + 1):
            yield run[i:i + n]


def mine_candidates(corpus: List[CorpusString], *, min_freq: int = 3,
                    max_n: int = 6) -> Dict[str, Candidate]:
    """Stage 1: high-recall Han n-gram mining over the full corpus,
    sentence interiors included. Keeps freq >= min_freq OR any standalone/
    enclosure evidence; prunes sub-grams dominated by a longer candidate."""
    cands: Dict[str, Candidate] = {}
    for s in corpus:
        seen_here = set()
        for run in _HAN_RUN_RE.findall(s.zh):
            for gram in _grams(run, max_n):
                if gram in seen_here:
                    continue
                if gram[0] in _EDGE_STOP or gram[-1] in _EDGE_STOP:
                    continue
                if all(ch in _EDGE_STOP for ch in gram):
                    continue
                seen_here.add(gram)
                c = cands.setdefault(gram, Candidate(zh=gram))
                c.freq += 1
                pos = s.zh.find(gram)
                before = s.zh[pos - 1] if pos > 0 else ""
                after = s.zh[pos + len(gram)] \
                    if pos + len(gram) < len(s.zh) else ""
                if before in _ENCLOSURE_OPEN and after in _ENCLOSURE_CLOSE:
                    c.enclosed += 1
                if len(c.examples) < 3:
                    c.examples.append((s.zh, s.en))
        if _HAN_RUN_RE.fullmatch(s.zh):            # whole string IS a term?
            c = cands.get(s.zh)
            if c is not None:
                c.standalone += 1
                for en, n in s.en_all.items():     # every distinct rendering
                    c.renderings[en] += n
                    if en[:1].isupper():
                        c.en_cap += n

    kept = {t: c for t, c in cands.items()
            if c.freq >= min_freq or c.standalone or c.enclosed}
    # sub-gram dominance pruning: 瘟疫 dies when 瘟疫点 has ~same support
    for t in sorted(kept, key=len):
        c = kept.get(t)
        if c is None:
            continue
        for u, cu in kept.items():
            if u != t and t in u and cu.freq >= 0.8 * c.freq:
                kept.pop(t, None)
                break
    return kept


# ------------------------------------------------------------------ stage 2

class TermVerdict(BaseModel):
    i: int
    keep: bool
    en: str = ""
    category: str = ""


class TermVerdicts(BaseModel):
    verdicts: List[TermVerdict]


_FILTER_SYSTEM = (
    "You are a game-localization terminologist. You receive candidate "
    "Chinese terms statistically mined from a game's string corpus. Keep "
    "a candidate only if it is a real GLOSSARY TERM: an item, skill, "
    "character, faction, place, resource, stat, or game-system name. Drop "
    "generic vocabulary, verb phrases, sentence fragments, and UI "
    "boilerplate. For kept terms, suggest an English rendering (short, "
    "1-4 words) and a category. Respond with ONLY JSON: "
    '{"verdicts": [{"i": <index>, "keep": true/false, '
    '"en": "<EN or empty>", "category": "<Item/Skill/Character/Place/'
    'System/Stat/Faction/Other>"}]} — one verdict per candidate, every '
    "index answered.")


def heuristic_keep(c: Candidate) -> bool:
    """Deterministic Stage-2 fallback: demand independent evidence beyond
    raw frequency, and term-plausible length."""
    return (len(c.zh) <= TERM_MAX_ZH_CHARS
            and (c.standalone > 0 or c.enclosed > 0 or c.freq >= 6))


def filter_candidates(cands: Dict[str, Candidate], provider=None, *,
                      batch: int = 40
                      ) -> Tuple[Dict[str, Candidate], List[str], str]:
    """Stage 2. Returns (kept, dropped_terms, mode). With a provider the
    LLM adjudicates in batches (keep/drop + EN suggestion, stored as
    ``llm_en``/``llm_category`` attributes); without one, heuristic_keep."""
    if provider is None:
        kept = {t: c for t, c in cands.items() if heuristic_keep(c)}
        return kept, sorted(set(cands) - set(kept)), "heuristic"

    from .llm import complete_json
    ordered = sorted(cands.values(), key=lambda c: -c.score)
    kept: Dict[str, Candidate] = {}
    dropped: List[str] = []
    for start in range(0, len(ordered), batch):
        chunk = ordered[start:start + batch]
        payload = [{"i": i, "zh": c.zh, "freq": c.freq,
                    "standalone": c.standalone,
                    "example": c.examples[0][0] if c.examples else ""}
                   for i, c in enumerate(chunk)]
        result = complete_json(
            provider, _FILTER_SYSTEM,
            json.dumps(payload, ensure_ascii=False), TermVerdicts,
            temperature=0.0, max_tokens=4000)
        verdicts = {v.i: v for v in result.verdicts}
        for i, c in enumerate(chunk):
            v = verdicts.get(i)
            if v is not None and v.keep:
                c.llm_en = v.en.strip()            # type: ignore[attr-defined]
                c.llm_category = v.category.strip()  # type: ignore[attr-defined]
                kept[c.zh] = c
            else:
                dropped.append(c.zh)
    return kept, dropped, f"llm:{provider.name}/{provider.model}"


# ------------------------------------------------------------------ stage 3

@dataclass
class Extraction:
    glossary: dict                 # pipeline T1 shape
    conflicts: List[dict]          # standalone renderings disagree (tie)
    violations: List[dict]         # per-TERM aggregated locked violations
    needs_en: List[dict]           # mined terms with no rendering anywhere
    dropped: List[str]             # stage-2 rejects
    stats: dict


def assemble(kept: Dict[str, Candidate], decisions: List[TermDecision],
             corpus: List[CorpusString], *, game: str = "",
             locale: str = "en", filter_mode: str = "",
             corpus_meta: Optional[dict] = None) -> Extraction:
    """Stage 3: candidates × locked decisions × corpus evidence →
    T1 glossary + per-term open questions."""
    terms: Dict[str, dict] = {}
    conflicts: List[dict] = []
    needs_en: List[dict] = []

    for decision in decisions:
        if decision.family:            # rules validate, never enter terms
            continue
        terms[decision.zh] = {
            "translation": _en_base(decision.en), "type": "decision",
            "locked": True, "evidence": decision.origin,
            **({"supersedes": decision.zh_old}
               if decision.zh_old else {})}

    for zh_term, c in sorted(kept.items()):
        if zh_term in terms:
            continue
        llm_en = getattr(c, "llm_en", "")
        evidence = (f"corpus ×{c.freq}"
                    + (f", standalone ×{c.standalone}" if c.standalone
                       else ""))
        if c.renderings:
            (top, top_n), *rest = c.renderings.most_common(2)
            if rest and rest[0][1] == top_n:
                conflicts.append({"zh": zh_term,
                                  "renderings": sorted(c.renderings),
                                  "freq": c.freq})
                continue
            terms[zh_term] = {"translation": top, "type": "mined",
                              "locked": False, "evidence": evidence}
        elif llm_en:
            terms[zh_term] = {"translation": llm_en, "type": "mined-llm",
                              "locked": False, "evidence": evidence,
                              **({"category": c.llm_category}
                                 if getattr(c, "llm_category", "") else {})}
        else:
            needs_en.append({"zh": zh_term, "freq": c.freq,
                             "example": c.examples[0][0]
                             if c.examples else ""})

    # per-TERM violation aggregation: locked term present in zh, decided
    # rendering absent from a translated EN — one record per term.
    violations: List[dict] = []
    for decision in decisions:
        bad_keys: List[str] = []
        total = 0
        for s in corpus:
            if decision.zh in s.zh or (decision.zh_old
                                       and decision.zh_old in s.zh):
                total += 1
                renderings = s.en_all or ({s.en: 1} if s.en else {})
                if renderings and any(not _en_has(decision.en, en)
                                      for en in renderings):
                    bad_keys.append(s.key)
        if bad_keys:
            violations.append({
                "zh": decision.zh, "expected_en": _en_base(decision.en),
                "origin": decision.origin, "strings_total": total,
                "strings_violating": len(bad_keys),
                "sample_keys": bad_keys[:5]})

    flagged = 0
    for zh_term, entry in terms.items():
        if entry["locked"]:
            continue
        for decision in decisions:
            contains = (decision.zh in zh_term if decision.family
                        else decision.zh != zh_term
                        and decision.zh in zh_term)
            if contains and not _en_has(decision.en,
                                        entry["translation"]):
                kind = "family rule" if decision.family else "locked"
                entry["check"] = (f"contains {kind} {decision.zh}="
                                  f"{_en_base(decision.en)}")
                flagged += 1
                break

    stats = {**(corpus_meta or {}),
             "filter_mode": filter_mode,
             "kept_candidates": len(kept),
             "terms_locked": sum(1 for t in terms.values() if t["locked"]),
             "terms_mined": sum(1 for t in terms.values()
                                if not t["locked"]),
             "conflicts": len(conflicts), "violations": len(violations),
             "needs_en": len(needs_en),
             "flagged_against_locked": flagged}
    family_rules = {d.zh: {"translation": _en_base(d.en),
                           "origin": d.origin}
                    for d in decisions if d.family}
    glossary = {
        "metadata": {"game": game, "locale": locale,
                     "built_from": "corpus extraction (stages 0-3)",
                     **stats,
                     **({"family_rules": family_rules}
                        if family_rules else {})},
        "terms": dict(sorted(terms.items(),
                             key=lambda kv: (not kv[1]["locked"], kv[0])))}
    return Extraction(glossary=glossary, conflicts=conflicts,
                      violations=violations, needs_en=needs_en,
                      dropped=[], stats=stats)


def extract_glossary(po_paths: Sequence[Path],
                     decisions: List[TermDecision], *, provider=None,
                     min_freq: int = 3, max_n: int = 6, game: str = "",
                     locale: str = "en") -> Extraction:
    """Stages 0–3, end to end."""
    corpus = load_corpus([Path(p) for p in po_paths])
    mined = mine_candidates(corpus, min_freq=min_freq, max_n=max_n)
    kept, dropped, mode = filter_candidates(mined, provider)
    result = assemble(
        kept, decisions, corpus, game=game, locale=locale,
        filter_mode=mode,
        corpus_meta={"corpus_strings": sum(s.count for s in corpus),
                     "corpus_unique": len(corpus),
                     "candidates_mined": len(mined)})
    result.dropped = dropped
    return result


# ------------------------------------------------------------------ outputs

def write_extract_review_xlsx(result: Extraction, path: Path) -> int:
    """Term-level Glossary PE form: ONLY open questions. Conflicts and
    per-term violations on the main sheet (short terms, one row per term);
    needs-EN terms on a separate work-queue sheet; string-level facts stay
    in the audit. Returns the main-sheet row count."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    headers = list(GLOSSARY_PE_FORM_HEADERS)
    rows: List[dict] = []
    for e in result.conflicts:
        rows.append({
            "TermID": e["zh"], "EntryType": "Conflict", "Source": e["zh"],
            "Alternatives": " | ".join(e["renderings"]),
            "Evidence": f"standalone renderings tie, corpus ×{e['freq']} — "
                        "pick one via Reject&Modification"})
    for e in result.violations:
        rows.append({
            "TermID": e["zh"], "EntryType": "Violation", "Source": e["zh"],
            "Target_Original": f"({e['strings_violating']}/"
                               f"{e['strings_total']} strings lack it)",
            "Target_Suggested": e["expected_en"],
            "Evidence": f"{e['origin']}; e.g. "
                        + ", ".join(e["sample_keys"][:3])})
    for zh_term, entry in result.glossary["terms"].items():
        if entry.get("check"):
            rows.append({
                "TermID": zh_term, "EntryType": "Violation",
                "Source": zh_term, "Target_Original": entry["translation"],
                "Evidence": "⚠ " + entry["check"]})

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Glossary PE"
    sheet.append(headers)
    fills = {"Conflict": "FFC7A0", "Violation": "FF9999"}
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="333F50")
        cell.alignment = Alignment(vertical="center")
    for number, row in enumerate(rows, 1):
        sheet.append([row.get(h, "") for h in headers])
        sheet.cell(row=number + 1,
                   column=headers.index("EntryType") + 1).fill = \
            PatternFill("solid", start_color=fills[row["EntryType"]])
    for col, options in FORM_DROPDOWNS["glossary"].items():
        validation = DataValidation(
            type="list", formula1='"' + ",".join(options) + '"',
            allow_blank=True, showDropDown=False)
        sheet.add_data_validation(validation)
        letter = sheet.cell(
            row=1, column=headers.index(col) + 1).column_letter
        validation.add(f"{letter}2:{letter}{max(len(rows), 1) + 1}")
    for name, width in (("Source", 24), ("Target_Original", 30),
                        ("Target_Suggested", 30), ("Alternatives", 40),
                        ("Evidence", 45), ("PE_Decision", 28),
                        ("PE_Modification", 30), ("PE_Note", 22),
                        ("PE_Query", 22)):
        letter = sheet.cell(
            row=1, column=headers.index(name) + 1).column_letter
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"

    if result.needs_en:
        queue = book.create_sheet("Needs EN")
        queue.append(["zh", "Corpus freq", "Example string", "EN"])
        for e in result.needs_en:
            queue.append([e["zh"], e["freq"], e["example"], ""])
        for letter, width in (("A", 24), ("C", 60), ("D", 30)):
            queue.column_dimensions[letter].width = width
        queue.freeze_panes = "A2"

    guide = book.create_sheet("说明 How to fill")
    for line in (
            "Term-level review: one row per TERM (never per string).",
            "Same PE form family as MTPE/LQA tables — fill PE_* only.",
            "",
            "Conflict = standalone renderings tie; pick the winner via "
            "Reject&Modification (PE_Modification).",
            "Violation = a locked ruling is missing from N strings "
            "(evidence lists samples) or a mined term contradicts a "
            "locked term (⚠). Accept Suggested Translation confirms the "
            "ruling should be enforced across those strings.",
            "Needs EN sheet = mined terms with no rendering anywhere in "
            "the corpus — a translation work queue, not adjudication.",
            "The full extracted glossary is in glossary_terms.xlsx/.json; "
            "string-level detail is in extract_audit.json."):
        guide.append([line])
    guide.column_dimensions["A"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return len(rows)


def write_extraction_outputs(result: Extraction, out_dir: Path) -> Path:
    """glossary_terms.{json,xlsx} + extract_review.xlsx + audit."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ties = [{"zh": e["zh"], "renderings": e["renderings"],
             "reason": "standalone renderings tie"}
            for e in result.conflicts]
    write_term_glossary(result.glossary, ties, out_dir)
    write_extract_review_xlsx(result, out_dir / "extract_review.xlsx")
    (out_dir / "extract_audit.json").write_text(json.dumps(
        {"stats": result.stats, "conflicts": result.conflicts,
         "violations": result.violations, "needs_en": result.needs_en,
         "dropped": result.dropped},
        ensure_ascii=False, indent=1), encoding="utf-8")
    lines = ["# Glossary extraction audit", "",
             "| Stage | Result |", "|---|---|"]
    s = result.stats
    lines += [f"| 0 intake | {s.get('corpus_strings')} strings, "
              f"{s.get('corpus_unique')} unique |",
              f"| 1 mining | {s.get('candidates_mined')} candidates |",
              f"| 2 filter ({s.get('filter_mode')}) | "
              f"{s.get('kept_candidates')} kept, "
              f"{len(result.dropped)} dropped |",
              f"| 3 assembly | {s.get('terms_locked')} locked + "
              f"{s.get('terms_mined')} mined terms; "
              f"{s.get('conflicts')} conflicts, {s.get('violations')} "
              f"violated locked terms, {s.get('needs_en')} need EN, "
              f"{s.get('flagged_against_locked')} flagged |"]
    md = out_dir / "extract_audit.md"
    md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return md

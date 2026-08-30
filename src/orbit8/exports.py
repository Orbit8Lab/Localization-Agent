"""Deterministic exporters to the pipeline's standard formats.

Two target shapes (the localization-pipeline contracts):

- flat source JSON  {"<key>": "<source text>"} — translation input
- bilingual JSONL   {"key", "source_language", "target_language",
                     "source_text", "target_text"} — LQA/MT input

The AGENT decides which files map to which shape (that is judgment); the
conversion itself is code. A UE .po carries its own bilingual pairs
(msgid = source, msgstr = target), so a single target-language .po is a
complete bilingual source.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List, Tuple

from .schemas import SourceString


def read_po_entries(path: Path) -> List[Tuple[str, str, str, str]]:
    """(key, msgid, msgstr, location) tuples, with continuation lines
    joined. ``location`` is the ``#:`` reference comment (UE exports carry
    the widget/asset path there) — "" when the file has none."""
    entries: List[Tuple[str, str, str, str]] = []
    key, msgid, msgstr, mode = None, [], [], None
    location, next_location = "", ""

    def flush():
        nonlocal key, msgid, msgstr, mode, location
        source = _unescape("".join(msgid))
        if source:
            entries.append((key or f"po:{len(entries)}", source,
                            _unescape("".join(msgstr)), location))
        key, msgid, msgstr, mode, location = None, [], [], None, ""

    for line in Path(path).read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line.startswith("#:"):
            next_location = line[2:].strip()
        elif line.startswith("msgctxt "):
            flush()
            key, mode = _unescape(line[8:].strip().strip('"')), "ctxt"
            location, next_location = next_location, ""
        elif line.startswith("msgid "):
            msgid, mode = [line[6:].strip().strip('"')], "id"
            if key is None:            # entry without msgctxt
                location, next_location = next_location, ""
        elif line.startswith("msgstr "):
            msgstr, mode = [line[7:].strip().strip('"')], "str"
        elif line.startswith('"'):
            if mode == "id":
                msgid.append(line.strip('"'))
            elif mode == "str":
                msgstr.append(line.strip('"'))
    flush()
    return entries


def _unescape(text: str) -> str:
    return (text.replace("\\r", "\r").replace("\\n", "\n")
            .replace("\\t", "\t").replace('\\"', '"')
            .replace("\\\\", "\\"))


def emit_flat_json(records: List[SourceString], path: Path) -> int:
    """Pipeline translation-input format; duplicate keys are an input bug
    and refuse loudly rather than silently overwrite."""
    flat: Dict[str, str] = {}
    for record in records:
        if record.key in flat and flat[record.key] != record.text:
            raise ValueError(f"duplicate key {record.key!r} with "
                             f"conflicting texts")
        flat[record.key] = record.text
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(flat, ensure_ascii=False, indent=2),
                    encoding="utf-8")
    return len(flat)


def emit_bilingual_jsonl(files: List[Path], path: Path, *,
                         source_lang: str, target_lang: str,
                         columns=None, fallback=None) -> Tuple[int, int]:
    """Pairs from target-language .po files (msgid→msgstr). Entries with
    empty msgstr are INCLUDED with an empty target — an untranslated
    string is exactly what a downstream LQA must flag, and an MT pass must
    fill; dropping them at export was a recall hole. Returns
    (written, empty_targets).

    Sanity guard: when msgstr overwhelmingly equals msgid, the file is a
    SOURCE-language export (e.g. English.po), not a translation — pairing
    it would produce useless en→en rows, so refuse with guidance."""
    non_po = [Path(f) for f in files if Path(f).suffix.lower() != ".po"]
    if non_po:
        if len(non_po) != len(files):
            raise ValueError("mixing .po with other formats is not "
                             "supported — convert them separately")
        if fallback is None:
            raise ValueError(
                f"bilingual export reads .po natively; got "
                f"{non_po[0].name!r}. Converting it needs the "
                f"adapter-writer — run without --dry-run.")
        # The WHOLE set goes to the adapter at once. Looping file-by-file
        # is what made a per-locale layout unreadable: shown one file, an
        # adapter finds a single text column and has nothing to pair it
        # with. What layout these files are in is the adapter's judgment
        # to make, and it can only make it having seen them together.
        context = (
            f"This job translates {source_lang} → {target_lang}. You are "
            f"extracting the {target_lang} translations for review.\n"
            f"Files given, in argv order: "
            + ", ".join(f"argv[{n + 1}]={p.name}"
                        for n, p in enumerate(non_po)))
        if columns:
            # Naming the columns turns a MULTI-language sheet into a
            # readable pair. A term list with four target columns fits
            # neither "bilingual file" nor "source + target files", and
            # without this the adapter has to guess which of four
            # translations is wanted.
            context += (
                f"\nThis file holds SEVERAL languages in one sheet. Use "
                f"the column headed {columns[0]!r} as the source and "
                f"{columns[1]!r} as the target; ignore all other language "
                f"columns. There is no id column in a term list — derive "
                f"a stable key from the source term itself.")
        pairs = list(fallback(non_po, context=context))
        if columns:
            _check_columns_were_honoured(non_po[0], columns, pairs)
        empty = sum(1 for _k, _s, target, _l in pairs if not target.strip())
        identical = sum(1 for _k, source, target, _l in pairs
                        if target.strip() and source.strip() == target.strip())
        return _write_pairs(pairs, path, source_lang=source_lang,
                            target_lang=target_lang, empty=empty,
                            identical=identical)

    pairs, empty, identical = [], 0, 0
    for file in files:
        for key, source, target, location in read_po_entries(Path(file)):
            if not target.strip():
                empty += 1
            elif source.strip() == target.strip():
                identical += 1
            pairs.append((key, source, target, location))
    return _write_pairs(pairs, path, source_lang=source_lang,
                        target_lang=target_lang, empty=empty,
                        identical=identical)


def _write_pairs(pairs, path: Path, *, source_lang: str, target_lang: str,
                 empty: int, identical: int) -> Tuple[int, int]:
    """Validate a pair set and write it. Shared by the .po path and the
    adapter path so neither can skip a guard the other enforces."""
    filled = len(pairs) - empty
    if pairs and not filled:
        # EVERY target empty means the input held ONE language, not two —
        # a per-locale export whose partner source file lives elsewhere.
        # Writing it anyway produced 400 rows of the target language
        # mislabelled as source, with nothing to review: shaped like valid
        # input, useless as one, and silent about it.
        raise ValueError(
            f"all {len(pairs)} rows have an empty target — the input holds "
            f"ONE language, so there is nothing to review. If the "
            f"translations live in a separate file from the source, pass "
            f"BOTH: the source file first, then the target file.")
    if filled and identical / filled > 0.8:
        raise ValueError(
            f"{identical}/{filled} pairs have identical source and "
            f"target — this is a source-language file; standardize it "
            f"with output=source_json instead")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        for key, source, target, location in pairs:
            row = {"key": key, "source_language": source_lang,
                   "target_language": target_lang, "source_text": source,
                   "target_text": target}
            if location:               # UE #: widget/asset path, when known
                row["location"] = location
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")
    return len(pairs), empty


def _check_columns_were_honoured(source_file: Path, columns, pairs) -> None:
    """Verify a generated adapter used the columns it was TOLD to use.

    The failure this catches is the dangerous kind: asked for
    English + 日本語 from a five-language sheet, a generated adapter wrote
    header-matching logic that silently fell through to "first three
    columns in order" and produced 简体中文 + 繁體中文 instead. The output
    was well formed — right row count, right keys, non-empty texts — so
    every schema check passed. Only comparing against the sheet reveals it.

    So compare: read the real columns and confirm the emitted text
    actually came from the ones requested. A mismatch is reported rather
    than written, because a wrong glossary is worse than no glossary — it
    becomes the standard every later string is judged against.
    """
    try:
        from openpyxl import load_workbook
        if source_file.suffix.lower() not in (".xlsx", ".xlsm"):
            return                      # only sheets carry named columns
        book = load_workbook(source_file, read_only=True, data_only=True)
        sheet = book[book.sheetnames[0]]
        rows = list(sheet.iter_rows(max_row=6, values_only=True))
        book.close()
    except Exception:
        return                          # unreadable here is not a verdict
    if not rows:
        return
    header = ["" if cell is None else str(cell).strip() for cell in rows[0]]
    try:
        want_source = header.index(str(columns[0]).strip())
        want_target = header.index(str(columns[1]).strip())
    except ValueError:
        missing = [c for c in columns if str(c).strip() not in header]
        raise ValueError(
            f"column(s) {missing} not found in {source_file.name}; "
            f"its headers are {header}")

    expected = {(str(row[want_source]).strip(), str(row[want_target]).strip())
                for row in rows[1:]
                if row[want_source] and row[want_target]}
    if not expected:
        return
    produced = {(source.strip(), target.strip())
                for _key, source, target, _loc in pairs[:len(rows)]}
    if not (expected & produced):
        raise ValueError(
            f"the generated adapter did not use the requested columns: "
            f"asked for {columns[0]!r} + {columns[1]!r}, but the output "
            f"does not match those columns' values (sample expected "
            f"{sorted(expected)[:2]}). Re-run to regenerate the adapter, "
            f"or convert the sheet to two single-language files.")

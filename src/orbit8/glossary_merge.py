"""Glossary round merge — filled review decisions + previous glossary.

The incremental-round design (decision ledger, not snapshot):

- The filled Glossary PE form is read back with the standard vocabulary
  (``standards.LQA_PE_DECISIONS``); ``Reject&Modification`` rows become NEW
  LOCKED decisions with round provenance. Incomplete rows (decision given,
  required column empty per ``standards.DECISION_REQUIRES``) are reported,
  never guessed.
- The previous glossary (draft xlsx: 分类/中文/英文/状态) is integrated with
  the precedence ladder **locked decision > draft-confirmed (已确认) >
  corpus-mined**. A draft term superseded by a ruling's rename chain
  is retired, not carried. Draft-vs-corpus disagreement
  keeps the human draft but stamps a ⚠ drift check — the corpus petitions,
  a human decides.
- Nothing is mutated in place; output is a new versioned T1 file plus a
  delta report (locked-from-review / carried / draft-agrees /
  draft-carried / drift / superseded / incomplete).
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from .glossary_update import _en_base, write_term_glossary
from .standards import (DECISION_REQUIRES, GLOSSARY_PE_FORM_HEADERS,
                        LQA_PE_DECISIONS)


# ------------------------------------------------------------ form read-back

def read_glossary_pe_form(path: Path) -> Tuple[List[dict], List[dict]]:
    """Read a filled Glossary PE form. Returns (rows, problems): rows are
    dicts keyed by the standard headers; problems are rows whose decision
    is unknown or misses its required companion column."""
    import openpyxl
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = book["Glossary PE"]
    header = [c.value for c in sheet[1]]
    missing = [h for h in GLOSSARY_PE_FORM_HEADERS if h not in header]
    if missing:
        raise ValueError(f"{path}: not a Glossary PE form — missing "
                         f"columns {missing}")
    rows: List[dict] = []
    problems: List[dict] = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        row = {h: (str(v).strip() if v is not None else "")
               for h, v in zip(header, raw)}
        if not any(row.values()):
            continue
        rows.append(row)
        decision = row["PE_Decision"]
        if not decision:
            continue                       # blank = undecided, carried
        if decision not in LQA_PE_DECISIONS:
            problems.append({**row, "problem":
                             f"unknown decision {decision!r}"})
            continue
        required = DECISION_REQUIRES.get(decision)
        if required and not row.get(required):
            problems.append({**row, "problem":
                             f"{decision} requires {required}"})
    return rows, problems


# ------------------------------------------------------------- draft parser

_STATUS_CONFIRMED = "已确认"


def load_draft_xlsx(path: Path, *,
                    sheet: str = "术语表 Glossary") -> List[dict]:
    """Parse the legacy draft glossary (分类/中文/英文/状态). Categories
    fill down; slash forms in the source cell register every segment as
    an alias of the same EN."""
    import openpyxl
    book = openpyxl.load_workbook(path, data_only=True)
    ws = book[sheet] if sheet in book.sheetnames else book.active
    entries: List[dict] = []
    category = ""
    for raw in ws.iter_rows(min_row=2, values_only=True):
        cat, zh, en, status = [(str(v).strip() if v is not None else "")
                               for v in (list(raw) + ["", "", "", ""])[:4]]
        if cat:
            category = cat
        if not zh or not en:
            continue
        segments = [s.strip() for s in re.split(r"[/／]", zh) if s.strip()]
        for i, segment in enumerate(segments):
            entries.append({"zh": segment, "en": en, "category": category,
                            "confirmed": status == _STATUS_CONFIRMED,
                            **({"alias_of": segments[0]} if i else {})})
    return entries


# ------------------------------------------------------------------- merge

def merge_round(current_t1: dict, review_rows: List[dict],
                draft_terms: List[dict], *, round_id: str,
                game: str = "", locale: str = "en") -> Tuple[dict, dict]:
    """Three-way merge: current extraction glossary × filled review
    decisions × previous draft. Returns (new T1 glossary, delta report)."""
    terms: Dict[str, dict] = {zh: dict(entry) for zh, entry
                              in current_t1.get("terms", {}).items()}
    delta = {"locked_from_review": [], "reconfirmed": [],
             "draft_agrees": [], "draft_carried": [], "draft_drift": [],
             "draft_superseded": [], "incomplete": [], "unchanged": 0}

    # 1. review decisions become locked entries (round provenance)
    for row in review_rows:
        zh = row["Source"] or row["TermID"]
        decision = row["PE_Decision"]
        if not zh or not decision:
            continue
        if decision == "Reject&Modification":
            if not row["PE_Modification"]:
                delta["incomplete"].append(
                    {"zh": zh, "problem": "PE_Modification empty"})
                continue
            terms[zh] = {"translation": row["PE_Modification"],
                         "type": "decision", "locked": True,
                         "evidence": f"PE review {round_id}",
                         **({"note": row["PE_Note"]}
                            if row.get("PE_Note") else {})}
            delta["locked_from_review"].append(
                {"zh": zh, "en": row["PE_Modification"]})
        elif decision == "Accept Suggested Translation":
            en = row["Target_Suggested"] or terms.get(zh, {}).get(
                "translation", "")
            if en:
                terms[zh] = {"translation": _en_base(en),
                             "type": "decision", "locked": True,
                             "evidence": f"PE review {round_id}"}
                delta["reconfirmed"].append({"zh": zh, "en": en})
        elif decision == "Reject&Keep-as-it-is":
            if zh in terms:
                terms[zh].setdefault("evidence", "")
                terms[zh]["evidence"] += f"; kept as-is {round_id}"
                delta["reconfirmed"].append(
                    {"zh": zh, "en": terms[zh]["translation"]})
        # Reject&Cannot Answer: stays open, nothing written

    # rename chains: draft terms retired by a ruling's supersedes
    superseded_zh = {entry.get("supersedes"): zh
                     for zh, entry in terms.items()
                     if entry.get("supersedes")}

    # 2. previous draft, under the precedence ladder
    for d in draft_terms:
        zh, en = d["zh"], d["en"]
        if zh in superseded_zh:
            delta["draft_superseded"].append(
                {"zh": zh, "en": en, "by": superseded_zh[zh]})
            continue
        hit = terms.get(zh)
        if hit is None:
            terms[zh] = {"translation": en, "type": "draft",
                         "locked": False,
                         "evidence": "draft 2026-07"
                                     + (" (已确认)" if d["confirmed"]
                                        else ""),
                         **({"category": d["category"]}
                            if d["category"] else {}),
                         **({"alias_of": d["alias_of"]}
                            if d.get("alias_of") else {})}
            delta["draft_carried"].append({"zh": zh, "en": en})
        elif _en_base(en).lower() == _en_base(
                hit["translation"]).lower():
            hit["evidence"] = (hit.get("evidence", "")
                               + "; draft agrees").lstrip("; ")
            if d["category"] and not hit.get("category"):
                hit["category"] = d["category"]
            delta["draft_agrees"].append({"zh": zh, "en": en})
        elif hit["locked"]:
            delta["draft_superseded"].append(
                {"zh": zh, "en": en, "by": zh,
                 "now": hit["translation"]})
        else:
            # human draft (esp. 已确认) outranks corpus-mined — but the
            # corpus petitions: keep the drift visible for re-decision.
            if d["confirmed"]:
                corpus_en = hit["translation"]
                terms[zh] = {**hit, "translation": en, "type": "draft",
                             "check": f"corpus drift: game renders as "
                                      f"{corpus_en!r}",
                             **({"category": d["category"]}
                                if d["category"] else {})}
                delta["draft_drift"].append(
                    {"zh": zh, "draft": en, "corpus": corpus_en,
                     "applied": en})
            else:
                hit["check"] = f"draft (unconfirmed) had {en!r}"
                delta["draft_drift"].append(
                    {"zh": zh, "draft": en,
                     "corpus": hit["translation"],
                     "applied": hit["translation"]})

    delta["unchanged"] = len(terms) - len(delta["locked_from_review"]) \
        - len(delta["reconfirmed"]) - len(delta["draft_carried"]) \
        - len(delta["draft_drift"])

    meta_in = current_t1.get("metadata", {})
    glossary = {
        "metadata": {
            "game": game or meta_in.get("game", ""),
            "locale": locale or meta_in.get("locale", "en"),
            **({"family_rules": meta_in["family_rules"]}
               if meta_in.get("family_rules") else {}),
            "round": round_id,
            "built_from": "extraction + PE review + draft 2026-07",
            "locked_terms": sum(1 for t in terms.values()
                                if t.get("locked")),
            "mined_terms": sum(1 for t in terms.values()
                               if not t.get("locked")),
            "drift_flags": len(delta["draft_drift"]),
        },
        "terms": dict(sorted(terms.items(),
                             key=lambda kv: (not kv[1].get("locked"),
                                             kv[0])))}
    return glossary, delta


def write_merge_outputs(glossary: dict, delta: dict,
                        out_dir: Path) -> Path:
    """glossary_terms.{json,xlsx} + merge_delta.{json,md}."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    write_term_glossary(glossary, [], out_dir)
    (out_dir / "merge_delta.json").write_text(
        json.dumps(delta, ensure_ascii=False, indent=1), encoding="utf-8")
    lines = ["# Glossary merge delta — "
             + glossary["metadata"].get("round", ""), "",
             "| Bucket | Count |", "|---|---|"]
    for bucket in ("locked_from_review", "reconfirmed", "draft_agrees",
                   "draft_carried", "draft_drift", "draft_superseded",
                   "incomplete"):
        lines.append(f"| {bucket} | {len(delta[bucket])} |")
    lines.append(f"| unchanged | {delta['unchanged']} |")
    for bucket in ("locked_from_review", "draft_drift",
                   "draft_superseded", "incomplete"):
        if delta[bucket]:
            lines += ["", f"## {bucket}", ""]
            for e in delta[bucket]:
                lines.append("- " + json.dumps(e, ensure_ascii=False))
    md = out_dir / "merge_delta.md"
    md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return md

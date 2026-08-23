"""Standard PE form emitter (docs/STANDARDS.md §4).

One writer for every form in the family — MTPE (§4.1) and LQA PE (§4.2);
the glossary variant (§4.3) has its own builder in glossary_update/
term_extract. Agent columns are filled from ``rows``; every PE_* column is
left empty with the standard dropdowns attached, so the returning document
parses deterministically.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Sequence

from .standards import (FORM_DROPDOWNS, LQA_PE_FORM_HEADERS,
                        MTPE_FORM_HEADERS, PE_FILLED_COLUMNS)

_HEADERS = {"mtpe": MTPE_FORM_HEADERS, "lqa": LQA_PE_FORM_HEADERS}
_TITLES = {"mtpe": "MTPE", "lqa": "LQA PE"}

_GUIDE = (
    "Fill ONLY the PE_* columns; agent columns must come back unchanged.",
    "",
    "PE_Decision (dropdown): see column options. "
    "Reject&Modification → final target in PE_Modification, how/why in "
    "PE_Note. Reject&Cannot Answer → reason/question in PE_Query "
    "(forwarded to the dev team).",
    "PE_Categorization (Accuracy/Terminology/Tone/Fluency/Technical) and "
    "PE_Severity (Critical/Major/Minor) are filled when rejecting.",
    "Blank PE_Decision = not yet reviewed.")


def emit_pe_form(path: Path, kind: str, rows: List[Dict[str, str]], *,
                 extra_columns: Sequence[str] = ()) -> int:
    """Write a standard PE form workbook. ``kind`` is ``mtpe`` or ``lqa``;
    ``rows`` carry the agent-filled columns (PE_* values are ignored —
    they belong to the post-editor). ``extra_columns`` appends
    agent-authored context columns (e.g. an LQA scan's "Findings")
    AFTER the standard block, so the standard shape is never altered —
    only extended. Returns the row count."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    if kind not in _HEADERS:
        raise ValueError(f"unknown form kind {kind!r}; use mtpe or lqa")
    headers = list(_HEADERS[kind])
    for column in extra_columns:
        if column in headers:
            raise ValueError(f"extra column {column!r} collides with the "
                             f"standard {kind} form")
        headers.append(column)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = _TITLES[kind]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="333F50")
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append([("" if h in PE_FILLED_COLUMNS else row.get(h, ""))
                      for h in headers])
    for col, options in FORM_DROPDOWNS[kind].items():
        validation = DataValidation(
            type="list", formula1='"' + ",".join(options) + '"',
            allow_blank=True, showDropDown=False)
        sheet.add_data_validation(validation)
        letter = sheet.cell(
            row=1, column=headers.index(col) + 1).column_letter
        validation.add(f"{letter}2:{letter}{max(len(rows), 1) + 1}")
    widths = {"StringID": 34, "Source": 45, "Target_MT": 45,
              "Findings": 60,
              "Target_Original": 45, "Target_Suggested": 45,
              "PE_Decision": 26, "PE_Modification": 45, "PE_Note": 24,
              "PE_Query": 24}
    for name, width in widths.items():
        if name in headers:
            letter = sheet.cell(
                row=1, column=headers.index(name) + 1).column_letter
            sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"

    guide = book.create_sheet("说明 How to fill")
    for line in _GUIDE:
        guide.append([line])
    guide.column_dimensions["A"].width = 100
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return len(rows)

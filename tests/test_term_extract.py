"""Stages 0–3 of corpus-first glossary extraction."""
import json
from pathlib import Path

import openpyxl
import pytest

from orbit8.glossary_update import TermDecision
from orbit8.term_extract import (Candidate, CorpusString, assemble,
                                 extract_glossary, filter_candidates,
                                 heuristic_keep, load_corpus,
                                 mine_candidates, strip_markup,
                                 write_extraction_outputs)

PO_HEADER = '''msgid ""
msgstr ""
"Project-Id-Version: test\\n"

'''


def _po(tmp_path: Path, entries) -> Path:
    lines = [PO_HEADER]
    for i, (zh, en) in enumerate(entries):
        lines.append(f'msgctxt ",K{i}"\nmsgid "{zh}"\nmsgstr "{en}"\n\n')
    path = tmp_path / "corpus.po"
    path.write_text("".join(lines), encoding="utf-8")
    return path


# 瘟疫点 appears inside 5 different sentences (never standalone);
# 秘火使徒 appears standalone twice with agreeing EN; 幻觉陷阱 standalone
# with conflicting EN; markup/placeholders sprinkled in.
ENTRIES = [
    ("每回合清理一个瘟疫点", "Clear one plague point per turn"),
    ("瘟疫点会扩散到相邻区域", "Plague points spread to nearby areas"),
    ("摧毁瘟疫点获得奖励", "Destroy a plague point for rewards"),
    ("瘟疫点数量达到{0}时游戏结束", "Game over at {0} plague points"),
    ("使用道具清除瘟疫点", "Use an item to remove the plague point"),
    ("秘火使徒", "Secret Fire Apostle"),
    ("击败「秘火使徒」以后解锁", "Unlocks after defeating Secret Fire Apostle"),
    ("幻觉陷阱", "Hallucination Trap"),
    ("幻觉陷阱", "Hallucinogenic Trap"),
    ("<b>点击</b>继续", "Click to continue"),
]


def test_strip_markup():
    assert strip_markup("瘟疫点数量达到{0}时") == "瘟疫点数量达到 时"
    assert strip_markup("<b>点击</b>继续 %s &amp; x") == "点击 继续 x"


def test_load_corpus_dedups_and_cleans(tmp_path: Path):
    corpus = load_corpus([_po(tmp_path, ENTRIES)])
    by_zh = {s.zh: s for s in corpus}
    assert by_zh["幻觉陷阱"].count == 2          # dedup FIRST
    assert "{0}" not in by_zh["瘟疫点数量达到 时游戏结束"].zh


def test_mine_candidates_finds_embedded_terms(tmp_path: Path):
    corpus = load_corpus([_po(tmp_path, ENTRIES)])
    cands = mine_candidates(corpus, min_freq=3)
    assert "瘟疫点" in cands                     # sentence-interior mining
    plague = cands["瘟疫点"]
    assert plague.freq == 5 and plague.standalone == 0
    # sub-gram pruned: 瘟疫 dominated by 瘟疫点
    assert "瘟疫" not in cands
    apostle = cands["秘火使徒"]
    assert apostle.standalone == 1 and apostle.enclosed == 1
    assert apostle.renderings["Secret Fire Apostle"] == 1
    # edge stopwords never begin/end a candidate
    assert not any(t[0] in "的了是" or t[-1] in "的了是" for t in cands)


def test_heuristic_filter_demands_evidence():
    weak = Candidate(zh="数量达到", freq=3)          # freq-only, no evidence
    assert not heuristic_keep(weak)
    strong = Candidate(zh="瘟疫点", freq=6)
    assert heuristic_keep(strong)
    named = Candidate(zh="秘火使徒", freq=2, standalone=1)
    assert heuristic_keep(named)


class FakeProvider:
    name, model = "fake", "test"

    def complete(self, system, user, *, temperature=0.3, max_tokens=2000):
        items = json.loads(user)
        verdicts = [{"i": it["i"], "keep": "瘟疫" in it["zh"],
                     "en": "Plague Node" if "瘟疫" in it["zh"] else "",
                     "category": "System"} for it in items]
        return json.dumps({"verdicts": verdicts}, ensure_ascii=False)


def test_llm_filter_batches_and_annotates():
    cands = {"瘟疫点": Candidate(zh="瘟疫点", freq=5),
             "数量达到": Candidate(zh="数量达到", freq=4)}
    kept, dropped, mode = filter_candidates(cands, FakeProvider())
    assert set(kept) == {"瘟疫点"} and dropped == ["数量达到"]
    assert kept["瘟疫点"].llm_en == "Plague Node"
    assert mode == "llm:fake/test"


def test_assemble_locked_conflicts_violations(tmp_path: Path):
    corpus = load_corpus([_po(tmp_path, ENTRIES)])
    kept = mine_candidates(corpus, min_freq=3)
    kept, _, _ = filter_candidates(kept, None)
    decisions = [TermDecision(zh="瘟疫点", en="Plague Node", origin="CL")]
    result = assemble(kept, decisions, corpus, game="测试")
    terms = result.glossary["terms"]
    # decision locked and first
    assert terms["瘟疫点"]["locked"] and next(iter(terms)) == "瘟疫点"
    # standalone-majority election, agreeing renderings
    assert terms["秘火使徒"]["translation"] == "Secret Fire Apostle"
    # tie → conflict, excluded from glossary
    assert "幻觉陷阱" not in terms
    assert result.conflicts[0]["zh"] == "幻觉陷阱"
    # violations aggregated per TERM: all 5 瘟疫点 strings lack the ruling
    assert len(result.violations) == 1
    v = result.violations[0]
    assert v["strings_violating"] == 5 and v["strings_total"] == 5
    assert len(v["sample_keys"]) == 5


def test_end_to_end_outputs(tmp_path: Path):
    po = _po(tmp_path, ENTRIES)
    decisions = [TermDecision(zh="瘟疫点", en="Plague Node", origin="CL")]
    result = extract_glossary([po], decisions, game="测试")
    out = tmp_path / "out"
    write_extraction_outputs(result, out)
    glossary = json.loads((out / "glossary_terms.json").read_text())
    assert set(glossary) == {"metadata", "terms"}    # T1 shape
    book = openpyxl.load_workbook(out / "extract_review.xlsx")
    sheet = book["Glossary PE"]
    rows = list(sheet.iter_rows(values_only=True))
    types = [r[1] for r in rows[1:]]
    # ONE violation row for 瘟疫点 (not 5), plus the 幻觉陷阱 conflict
    assert types.count("Violation") >= 1
    assert types.count("Conflict") == 1
    sources = [r[2] for r in rows[1:]]
    assert sources.count("瘟疫点") == 1              # per-term aggregation
    assert all(len(str(s)) <= 10 for s in sources)   # short terms only
    assert "说明 How to fill" in book.sheetnames
    assert (out / "extract_audit.md").exists()

"""Glossary refresh from PE results: dedup-index join, conflict
resolution (decision > majority > review), term audit, suggestion
mining, and the decisions-xlsx loader."""
import json
from pathlib import Path

import openpyxl

from orbit8.glossary_update import (TermDecision, load_decisions_xlsx,
                                    refresh_glossary, write_update_outputs)

PE_PO = (
    'msgctxt ",K1A"\nmsgid "回血药水"\nmsgstr "Health Potion"\n\n'
    'msgctxt ",K1B"\nmsgid "回血药水"\nmsgstr "Health Potion"\n\n'
    # K2: duplicates disagree; decision 精神值=Spirit picks the compliant one
    'msgctxt ",K2A"\nmsgid "恢复精神值的药水"\n'
    'msgstr "Potion that restores Spirit"\n\n'
    'msgctxt ",K2B"\nmsgid "恢复精神值的药水"\n'
    'msgstr "Restore energy"\n\n'
    # K3: disagreeing duplicates, no decision applies, majority 2:1
    'msgctxt ",K3A"\nmsgid "金币"\nmsgstr "Gold Coin"\n\n'
    'msgctxt ",K3B"\nmsgid "金币"\nmsgstr "Gold Coin"\n\n'
    'msgctxt ",K3C"\nmsgid "金币"\nmsgstr "Golden coin"\n\n'
    # K4: 50/50 tie, no decision → open conflict
    'msgctxt ",K4A"\nmsgid "魔晶"\nmsgstr "Magic Crystal"\n\n'
    'msgctxt ",K4B"\nmsgid "魔晶"\nmsgstr "Mana Crystal"\n\n'
    # K5: agrees but violates the 瘟疫点 decision
    'msgctxt ",K5A"\nmsgid "净化瘟疫点"\nmsgstr "Purify the plague point"\n\n'
    'msgctxt ",K6A"\nmsgid "草药"\nmsgstr "Herbs"\n')


def _assets(tmp_path: Path) -> Path:
    d = tmp_path / "assets"
    d.mkdir()
    zh = {"asset_0": "回血药水", "asset_1": "恢复精神值的药水",
          "asset_2": "金币", "asset_3": "魔晶", "asset_4": "净化瘟疫点",
          "asset_5": "草药", "asset_6": "永不出现的条目"}
    en = {"asset_0": "Blood potion", "asset_1": "Restore energy",
          "asset_2": "Gold coin", "asset_3": "Magic Crystal",
          "asset_4": "Purify the plague point", "asset_5": "Herbs",
          "asset_6": "Never appears"}
    dedup = {"asset": {
        "asset_0": {"keys": ["K1A", "K1B"]},
        "asset_1": {"keys": ["K2A", "K2B"]},
        "asset_2": {"keys": ["K3A", "K3B", "K3C"]},
        "asset_3": {"keys": ["K4A", "K4B"]},
        "asset_4": {"keys": ["K5A"]},
        "asset_5": {"keys": ["K6A"]},
        "asset_6": {"keys": ["KGONE"]}}}
    (d / "zh_asset.json").write_text(
        json.dumps(zh, ensure_ascii=False, indent=2), encoding="utf-8")
    (d / "en_asset.json").write_text(
        json.dumps(en, ensure_ascii=False, indent=2), encoding="utf-8")
    (d / "dedup_index.json").write_text(
        json.dumps(dedup, ensure_ascii=False), encoding="utf-8")
    return d


def test_refresh_join_conflicts_audit(tmp_path: Path):
    assets = _assets(tmp_path)
    pe = tmp_path / "Game_en.po"
    pe.write_text(PE_PO, encoding="utf-8")
    decisions = [TermDecision(zh="精神值", en="Spirit", origin="user"),
                 TermDecision(zh="瘟疫点", en="Plague Node",
                              origin="CL-02")]
    result = refresh_glossary(assets, pe, decisions)
    assert result.counts() == {
        "updated": 1,                # asset_0 Blood potion→Health Potion
        "conflicts_resolved": 2,     # asset_1 (decision), asset_2 (majority)
        "conflicts_open": 1,         # asset_3 tie
        "unresolved": 1,             # asset_6 key not in PE po
        "unchanged": 2,              # asset_4 agrees-with-old, asset_5
        "term_violations": 1,        # asset_4 lacks 'Plague Node'
        "suggested_terms": 0}
    assert result.updated_en["asset_0"] == "Health Potion"
    assert result.updated_en["asset_1"] == "Potion that restores Spirit"
    by_asset = {e["asset"]: e for e in result.conflicts_resolved}
    assert by_asset["asset_1"]["resolved_by"] == "decision"
    assert by_asset["asset_2"]["resolved_by"] == "majority"
    assert result.updated_en["asset_2"] == "Gold Coin"
    assert result.updated_en["asset_3"] == "Magic Crystal"   # kept old
    violation = result.term_violations[0]
    assert (violation["asset"], violation["expected_en"]) == (
        "asset_4", "Plague Node")


def test_suggestion_mining(tmp_path: Path):
    """Repeated old→new replacements without a covering decision are
    surfaced; anything a decision already covers is not."""
    assets = tmp_path / "assets"
    assets.mkdir()
    n = 4
    zh = {f"asset_{i}": f"马车提示{i}" for i in range(n)}
    en = {f"asset_{i}": f"Protect the carriage {i}" for i in range(n)}
    dedup = {"asset": {f"asset_{i}": {"keys": [f"K{i}"]}
                       for i in range(n)}}
    po = "".join(
        f'msgctxt ",K{i}"\nmsgid "马车提示{i}"\n'
        f'msgstr "Protect the convoy {i}"\n\n' for i in range(n))
    (assets / "zh_asset.json").write_text(
        json.dumps(zh, ensure_ascii=False), encoding="utf-8")
    (assets / "en_asset.json").write_text(
        json.dumps(en, ensure_ascii=False), encoding="utf-8")
    (assets / "dedup_index.json").write_text(
        json.dumps(dedup), encoding="utf-8")
    pe = tmp_path / "pe.po"
    pe.write_text(po, encoding="utf-8")

    mined = refresh_glossary(assets, pe, [])
    assert mined.suggestions == [
        {"old_en": "carriage", "new_en": "convoy", "occurrences": 4}]
    covered = refresh_glossary(
        assets, pe, [TermDecision(zh="马车", en="Convoy")])
    assert covered.suggestions == []


def test_plural_gloss_notation(tmp_path: Path):
    """Workbook finals like 'Apostle(s)' must accept both Apostle and
    Apostles — no false violations from the (s) notation."""
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "zh_asset.json").write_text(json.dumps(
        {"asset_0": "秘火使徒获胜"}, ensure_ascii=False), encoding="utf-8")
    (assets / "en_asset.json").write_text(json.dumps(
        {"asset_0": "Secret Fire Apostles win"}), encoding="utf-8")
    (assets / "dedup_index.json").write_text(json.dumps(
        {"asset": {"asset_0": {"keys": ["K0"]}}}), encoding="utf-8")
    pe = tmp_path / "pe.po"
    pe.write_text('msgctxt ",K0"\nmsgid "秘火使徒获胜"\n'
                  'msgstr "Secret Fire Apostles win"\n', encoding="utf-8")
    result = refresh_glossary(assets, pe, [TermDecision(
        zh="秘火使徒", en="Secret Fire Apostle(s)", origin="J-01")])
    assert result.term_violations == []


def test_load_decisions_xlsx(tmp_path: Path):
    book = openpyxl.Workbook()
    mapping = book.active
    mapping.title = "内部·映射留档"
    mapping.append(["映射ID", "旧源文(zh)", "新源文(zh, dev终版)",
                    "EN定稿", "映射类型", "源文修改状态", "备注"])
    mapping.append(["M-05", "SAN值", "精神值", "Spirit", "命名", "", ""])
    mapping.append(["M-04", "马车", "车队", "Convoy", "命名", "", ""])
    mapping.append(["M-xx", "", "", "", "", "", ""])       # blank: skipped
    ruling = book.create_sheet("内部·翻译裁定")
    ruling.append(["裁定ID", "议题", "选项", "推荐及理由", "影响PO条目数",
                   "裁定结果(填这里)", "裁定后动作"])
    ruling.append(["J-03", "铁匠铺EN", "A/B/C", "…", 13,
                   "B → Blacksmith", "…"])
    ruling.append(["J-02", "雾月岛/雾月城EN", "A/B/C", "…", 6,
                   "A → Misty Moon Island / Misty Moon City", "…"])
    ruling.append(["J-09", "无法解析的议题", "A/B", "…", 1, "还没裁定", "…"])
    ruling.append(["J-05", "致幻族EN", "A/B", "…", 9,
                   "A → Hallucination 族", "feed TermBase→全族对齐"])
    path = tmp_path / "decisions.xlsx"
    book.save(path)

    decisions = load_decisions_xlsx(path)
    as_tuples = {(d.zh, d.en, d.zh_old) for d in decisions}
    assert ("精神值", "Spirit", "SAN值") in as_tuples
    assert ("车队", "Convoy", "马车") in as_tuples
    assert ("铁匠铺", "Blacksmith", None) in as_tuples
    assert ("雾月岛", "Misty Moon Island", None) in as_tuples
    assert ("雾月城", "Misty Moon City", None) in as_tuples
    # family ruling: 族 suffix is shorthand, stripped from BOTH sides —
    # a RULE over every term containing 致幻, never a term itself
    family = next(d for d in decisions if d.zh == "致幻")
    assert family.en == "Hallucination" and family.family is True
    assert "族/family" in family.origin
    assert not any(d.zh == "致幻族" for d in decisions)
    assert all(d.family is False for d in decisions if d.zh != "致幻")
    assert len(decisions) == 6                 # unparseable rows skipped

    # distill: family rules go to metadata, not terms; members that lack
    # the rendering get flagged
    from orbit8.glossary_update import distill_term_glossary
    glossary, _ = distill_term_glossary(
        {"a1": "致幻蘑菇", "a2": "致幻陷阱"},
        {"a1": "Hallucination Mushroom", "a2": "Illusion Trap"},
        decisions)
    assert "致幻" not in glossary["terms"]
    assert "致幻族" not in glossary["terms"]
    assert glossary["metadata"]["family_rules"]["致幻"][
        "translation"] == "Hallucination"
    assert "check" not in glossary["terms"]["致幻蘑菇"]
    assert "family rule" in glossary["terms"]["致幻陷阱"]["check"]


def test_review_xlsx_rows_and_dropdown(tmp_path: Path):
    from orbit8.glossary_update import write_review_xlsx
    from orbit8.standards import (GLOSSARY_PE_FORM_HEADERS,
                                  LQA_PE_DECISIONS)
    headers = list(GLOSSARY_PE_FORM_HEADERS)
    assets = _assets(tmp_path)
    pe = tmp_path / "pe.po"
    pe.write_text(PE_PO, encoding="utf-8")
    result = refresh_glossary(
        assets, pe, [TermDecision(zh="瘟疫点", en="Plague Node",
                                  origin="CL-02")])
    path = tmp_path / "review.xlsx"
    count = write_review_xlsx(result, path)
    book = openpyxl.load_workbook(path)
    sheet = book["Glossary PE"]
    rows = list(sheet.iter_rows(values_only=True))
    assert list(rows[0]) == headers
    assert count == len(rows) - 1
    types = [r[headers.index("EntryType")] for r in rows[1:]]
    # review-first ordering: conflicts before violations before updates
    assert types.index("Conflict") < types.index("Violation")
    assert types.index("Violation") < types.index("Update")
    by_type = {t: types.count(t) for t in set(types)}
    # without a Spirit ruling, asset_1's 50/50 tie is open too
    assert by_type["Conflict"] == 2        # asset_1 + asset_3 ties
    assert by_type["Violation"] == 1       # asset_4 Plague Node
    assert by_type["Update"] == 2          # asset_0 + asset_2 majority
    alternatives = " ; ".join(
        str(r[headers.index("Alternatives")])
        for r in rows[1:] if r[headers.index("EntryType")] == "Conflict")
    assert "Magic Crystal" in alternatives          # asset_3's tie listed
    # PE_Decision carries the standard LQA decision dropdown
    validations = sheet.data_validations.dataValidation
    assert any(LQA_PE_DECISIONS[1] in (dv.formula1 or "")
               for dv in validations)
    assert "说明 How to fill" in book.sheetnames


def test_distill_term_glossary(tmp_path: Path):
    from orbit8.glossary_update import (distill_term_glossary,
                                        is_term_like,
                                        write_term_glossary)
    # term-likeness: short names yes; sentences/的-phrases/long EN no
    assert is_term_like("回血药水", "Health Potion")
    assert is_term_like("太刀", "Tachi")
    assert not is_term_like("恢复精神值的药水", "Potion")        # 的-phrase
    assert not is_term_like("夜晚时精神值下降，请注意。", "x")   # sentence
    assert not is_term_like("金币", "a very long english rendering here")
    assert not is_term_like("%", "%")                          # no CJK
    assert not is_term_like("(0/3)", "(0/3)")
    assert not is_term_like("1号玩家", "Player 1")             # digit-led

    zh = {"a0": "回血药水", "a1": "回血药水", "a2": "回血药水",
          "a3": "金币", "a4": "金币",
          "a5": "恢复精神值的药水", "a6": "太刀", "a7": "玩家太刀"}
    en = {"a0": "Health Potion", "a1": "Health Potion",
          "a2": "Blood potion",                    # minority → outvoted
          "a3": "Gold Coin", "a4": "Golden coin",  # 1:1 tie → excluded
          "a5": "Potion that restores Spirit", "a6": "Longsword",
          "a7": "Player Greatsword"}   # contains locked 太刀, no Tachi
    decisions = [TermDecision(zh="太刀", en="Tachi", origin="J-04"),
                 TermDecision(zh="精神值", en="Spirit(s)", origin="M-05",
                              zh_old="SAN值")]
    glossary, ties = distill_term_glossary(zh, en, decisions, game="G")
    terms = glossary["terms"]
    assert terms["回血药水"] == {"translation": "Health Potion",
                             "type": "mined", "locked": False,
                             "evidence": "PE ×3"}
    assert "金币" not in terms
    assert ties == [{"zh": "金币", "renderings": ["Gold Coin",
                                                "Golden coin"],
                     "reason": "tie between PE renderings"}]
    # decision beats the mined 'Longsword'; (s) notation normalized
    assert terms["太刀"]["translation"] == "Tachi"
    assert terms["太刀"]["locked"] is True
    assert terms["精神值"] == {"translation": "Spirit", "type": "decision",
                            "locked": True, "evidence": "M-05",
                            "supersedes": "SAN值"}
    # locked entries sort first
    assert list(terms)[:2] == ["太刀", "精神值"]
    assert glossary["metadata"]["locked_terms"] == 2
    # mined term containing a locked term without its rendering is flagged
    assert terms["玩家太刀"]["check"] == "contains locked 太刀=Tachi"
    assert glossary["metadata"]["flagged_against_locked"] == 1

    path = write_term_glossary(glossary, ties, tmp_path / "out")
    book = openpyxl.load_workbook(path)
    rows = list(book["Glossary"].iter_rows(values_only=True))
    assert rows[0] == ("zh", "EN", "Locked", "Source", "Notes")
    assert ("太刀", "Tachi", "LOCKED", "J-04", None) in rows
    assert ("精神值", "Spirit", "LOCKED", "M-05",
            "supersedes SAN值") in rows
    assert "Ties (excluded)" in book.sheetnames
    saved = json.loads(
        (tmp_path / "out" / "glossary_terms.json").read_text("utf-8"))
    assert saved["terms"]["回血药水"]["translation"] == "Health Potion"


def test_outputs_preserve_format(tmp_path: Path):
    assets = _assets(tmp_path)
    pe = tmp_path / "pe.po"
    pe.write_text(PE_PO, encoding="utf-8")
    result = refresh_glossary(assets, pe, [])
    md = write_update_outputs(result, tmp_path / "out", assets)
    updated = json.loads(
        (tmp_path / "out" / "en_asset.json").read_text("utf-8"))
    assert list(updated) == [f"asset_{i}" for i in range(7)]  # order kept
    original = json.loads((assets / "en_asset.json").read_text("utf-8"))
    assert original["asset_0"] == "Blood potion"     # source untouched
    assert md.exists()
    assert (tmp_path / "out" / "glossary_update_audit.json").exists()

"""External LQA as a real side entrance to the pipeline.

`orbit8 lqa run` audits translations SOMEBODY ELSE produced. It reads the
intake and a pairs file and nothing else — no ingest, no context, no
glossary lock, no translation. That is the design, and two things quietly
prevented it:

1. `emit_bilingual_jsonl` hard-rejected anything but `.po`, so a client
   sending an xlsx/csv of translations for review had no way in — while
   the identical file as a SOURCE was handled fine by the Adapter-Writer.
   An asymmetry, not a decision.
2. `run_external_lqa` read the s2 style brief unconditionally, an artifact
   only CONTEXT produces — a stage an external audit never runs.

Together they made a documented entry point unreachable.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

from orbit8.controller import Job
from orbit8.exports import emit_bilingual_jsonl
from orbit8.schemas import IntakeBrief


@pytest.fixture
def job(tmp_path) -> Job:
    return Job.init(tmp_path / "jobs", "audit",
                    intake=IntakeBrief(game="G", source_lang="en",
                                       target_locales=["zh-CN"]),
                    source_files=[])


def _po(path: Path, pairs: list) -> Path:
    path.write_text(
        "\n\n".join(f'msgctxt ",{k}"\nmsgid "{s}"\nmsgstr "{t}"'
                    for k, s, t in pairs) + "\n", encoding="utf-8")
    return path


# ------------------------------------------------ the style-brief blocker

def test_the_style_brief_is_optional_for_an_external_audit(job):
    """An external audit enters at S5, so no s2 artifact exists. The LQA
    context already types style_brief as Optional and the cascade handles
    None — the hard read was the only thing making the entry point
    impossible."""
    assert job._style_or_none() is None


def test_the_strict_reader_still_raises_where_it_should(job):
    """Only the external path tolerates a missing brief. A normal
    lifecycle run reaching CONTEXT-dependent work with no brief is a real
    error and must stay loud."""
    from orbit8.store import ArtifactError
    with pytest.raises(ArtifactError):
        job._style()


# ------------------------------------------- bilingual formats beyond .po

def test_a_po_still_works_without_any_fallback(tmp_path):
    """The native path must not regress: .po needs no adapter."""
    source = _po(tmp_path / "t.po",
                 [("A", "Start", "开始"), ("B", "Quit", "退出")])
    written, empty = emit_bilingual_jsonl(
        [source], tmp_path / "out.jsonl",
        source_lang="en", target_lang="zh-CN")
    assert written == 2 and empty == 0


def test_a_non_po_without_a_fallback_explains_itself(tmp_path):
    """Dry-run cannot generate an adapter, so the refusal has to say what
    to do rather than just naming the suffix."""
    csv = tmp_path / "pairs.csv"
    csv.write_text("Key,English,Chinese\nA,Start,开始\n", encoding="utf-8")
    with pytest.raises(ValueError) as excinfo:
        emit_bilingual_jsonl([csv], tmp_path / "out.jsonl",
                             source_lang="en", target_lang="zh-CN")
    assert "adapter-writer" in str(excinfo.value)


def test_a_non_po_goes_through_the_fallback(tmp_path):
    """The fix: any format the adapter-writer can read is a valid
    bilingual input, exactly as it already is for source ingest."""
    csv = tmp_path / "pairs.csv"
    csv.write_text("Key,English,Chinese\nA,Start,开始\nB,Quit,退出\n",
                   encoding="utf-8")

    def fake_adapter(paths, *, context="", **kw):
        assert isinstance(paths, list)      # the WHOLE set, not one file
        assert "en" in context and "zh-CN" in context
        return [("A", "Start", "开始", str(paths[0])),
                ("B", "Quit", "退出", str(paths[0]))]

    written, empty = emit_bilingual_jsonl(
        [csv], tmp_path / "out.jsonl", source_lang="en",
        target_lang="zh-CN", fallback=fake_adapter)
    assert written == 2 and empty == 0

    rows = [json.loads(line) for line in
            (tmp_path / "out.jsonl").read_text(encoding="utf-8").splitlines()]
    assert rows[0]["source_text"] == "Start"
    assert rows[0]["target_text"] == "开始"


def test_an_untranslated_row_survives_alongside_translated_ones(tmp_path):
    """An untranslated string is exactly what a review must flag —
    dropping it at export is a recall hole, and that rule has to hold on
    the adapter path too."""
    csv = tmp_path / "pairs.csv"
    csv.write_text("Key,English,Chinese\nA,Start,开始\nB,Quit,\n",
                   encoding="utf-8")

    written, empty = emit_bilingual_jsonl(
        [csv], tmp_path / "out.jsonl", source_lang="en",
        target_lang="zh-CN",
        fallback=lambda p, *, context="", **kw: [
            ("A", "Start", "开始", str(p[0])),
            ("B", "Quit", "", str(p[0]))])
    assert written == 2 and empty == 1


def test_an_export_with_NO_translations_is_refused(tmp_path):
    """The bug that produced 400 unusable rows: a per-locale file holds
    ONE language, so an adapter shown it alone emits that language as
    `source` with every target empty. Correctly shaped, nothing to
    review, and silent — an LQA run would have found no defects because
    there was no translation to inspect."""
    csv = tmp_path / "ja_only.csv"
    csv.write_text("Key,Japanese\nA,ゲーム開始\n", encoding="utf-8")

    with pytest.raises(ValueError) as excinfo:
        emit_bilingual_jsonl(
            [csv], tmp_path / "out.jsonl", source_lang="en",
            target_lang="ja",
            fallback=lambda p, *, context="", **kw: [
                (f"K{n}", "ゲーム開始", "", str(p[0])) for n in range(5)])
    message = str(excinfo.value)
    assert "empty target" in message
    assert "pass BOTH" in message           # and how to fix it


def test_two_files_are_handed_to_the_adapter_together(tmp_path):
    """THE fix. A per-locale layout is only legible when the adapter sees
    the source export and the target export at once — shown one at a
    time it can find a source with no target and nothing else."""
    source = tmp_path / "Game (Source).xlsx"
    target = tmp_path / "Game_ja.xlsx"
    for path in (source, target):
        path.write_bytes(b"PK\x03\x04stub")

    seen = {}

    def adapter(paths, *, context="", **kw):
        seen["files"] = [p.name for p in paths]
        seen["context"] = context
        return [("A", "Start Game", "ゲーム開始", str(paths[0]))]

    written, _empty = emit_bilingual_jsonl(
        [source, target], tmp_path / "out.jsonl", source_lang="en",
        target_lang="ja", fallback=adapter)

    assert written == 1
    assert seen["files"] == ["Game (Source).xlsx", "Game_ja.xlsx"]
    # the context names the languages AND the argv positions, which is
    # what lets the adapter tell which file holds which language
    assert "en → ja" in seen["context"]
    assert "argv[1]=Game (Source).xlsx" in seen["context"]


def test_a_source_language_file_is_still_refused(tmp_path):
    """The sanity guard must survive the new path: a file whose 'targets'
    equal its sources is a SOURCE export, and pairing it produces useless
    en→en rows."""
    csv = tmp_path / "english.csv"
    csv.write_text("Key,A,B\n", encoding="utf-8")
    with pytest.raises(ValueError) as excinfo:
        emit_bilingual_jsonl(
            [csv], tmp_path / "out.jsonl", source_lang="en",
            target_lang="zh-CN",
            fallback=lambda p, *, context="", **kw: [
                (f"K{n}", "Start", "Start", str(p[0])) for n in range(10)])
    assert "source-language file" in str(excinfo.value)


# ------------------------------------------- inspecting a spreadsheet

def _chat(tmp_path):
    from orbit8.llm import EchoProvider
    from orbit8.orchestrator import ChatOrchestrator
    job = Job.init(tmp_path / "proj" / "jobs", "j",
                   intake=IntakeBrief(game="G", source_lang="en",
                                      target_locales=["ja"]),
                   source_files=[])
    return ChatOrchestrator(job, EchoProvider("ja"), operator="t",
                            dry_run=True)


def test_inspecting_an_xlsx_shows_columns_not_zip_bytes(tmp_path):
    """An xlsx is a zip. Text-peeking it handed the model `PK\\x03\\x04…`,
    which answers nothing about the columns — so it called inspect again
    and again and burned the whole step budget on a tool that "succeeded"
    every time."""
    openpyxl = pytest.importorskip("openpyxl")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Strings"
    sheet.append(["Key", "English", "Japanese"])
    sheet.append(["UI_START", "Start Game", "ゲーム開始"])
    path = tmp_path / "proj" / "10-received"
    path.mkdir(parents=True)
    book.save(path / "Bilingual.xlsx")

    result = json.loads(_chat(tmp_path)._t_inspect_file(
        {"path": "10-received/Bilingual.xlsx"}))
    assert "PK" not in str(result.get("head", ""))
    preview = result["sheet_preview"][0]
    assert preview["header"] == ["Key", "English", "Japanese"]
    assert preview["sample_rows"][0][2] == "ゲーム開始"


def test_inspecting_a_csv_shows_its_header(tmp_path):
    path = tmp_path / "proj" / "10-received"
    path.mkdir(parents=True)
    (path / "pairs.csv").write_text(
        "Key,English,Japanese\nUI_START,Start Game,ゲーム開始\n",
        encoding="utf-8")

    result = json.loads(_chat(tmp_path)._t_inspect_file(
        {"path": "10-received/pairs.csv"}))
    assert result["header"] == ["Key", "English", "Japanese"]


def test_an_unreadable_spreadsheet_reports_the_error(tmp_path):
    """Never fall back to a byte peek: binary noise is what caused the
    loop this exists to prevent."""
    path = tmp_path / "proj" / "10-received"
    path.mkdir(parents=True)
    (path / "broken.xlsx").write_bytes(b"not really a spreadsheet")

    result = json.loads(_chat(tmp_path)._t_inspect_file(
        {"path": "10-received/broken.xlsx"}))
    assert "error" in result
    assert "head" not in result


def test_a_two_column_file_is_refused_before_any_model_call(tmp_path):
    """The structural check that turned 190 seconds into one. A single
    non-.po file with too few columns cannot hold a source AND a target,
    so the adapter-writer would spend three attempts discovering that a
    file cannot be paired with itself."""
    pytest.importorskip("openpyxl")
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Key", "Japanese"])
    sheet.append(["A", "ゲーム開始"])
    received = tmp_path / "proj" / "10-received"
    received.mkdir(parents=True)
    book.save(received / "ja_only.xlsx")

    chat = _chat(tmp_path)
    result = chat._t_standardize({"files": ["10-received/ja_only.xlsx"],
                                  "output": "bilingual_jsonl",
                                  "target_lang": "ja"})
    assert "too few to hold both" in result
    assert "pass BOTH" in result


def test_a_three_column_file_is_not_pre_refused(tmp_path):
    """The check must not block a genuine bilingual sheet — it is a
    structural impossibility test, not a quality one."""
    pytest.importorskip("openpyxl")
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Key", "English", "Japanese"])
    sheet.append(["A", "Start", "開始"])
    received = tmp_path / "proj" / "10-received"
    received.mkdir(parents=True)
    book.save(received / "both.xlsx")

    # dry_run stops at the adapter-writer, which is PAST the column check:
    # reaching that error proves the structural test let this file through.
    with pytest.raises(ValueError) as excinfo:
        _chat(tmp_path)._t_standardize({"files": ["10-received/both.xlsx"],
                                        "output": "bilingual_jsonl",
                                        "target_lang": "ja"})
    assert "dry-run cannot generate an adapter" in str(excinfo.value)


# -------------------------------------- multi-language glossary sheets

def _glossary(tmp_path: Path) -> Path:
    """A five-language term sheet — neither pipeline format fits it."""
    pytest.importorskip("openpyxl")
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["English", "简体中文", "繁體中文", "한국어", "日本語"])
    sheet.append(["Barrier", "屏障", "屏障", "방벽", "バリア"])
    sheet.append(["Season Trial", "季节试炼", "季節試煉", "계절 시련",
                  "季節の試練"])
    received = tmp_path / "proj" / "10-received"
    received.mkdir(parents=True, exist_ok=True)
    path = received / "Glossary.xlsx"
    book.save(path)
    return path


def test_a_wrong_column_choice_is_caught_before_writing(tmp_path):
    """The dangerous failure: asked for English + 日本語, a generated
    adapter fell through to "first three columns" and produced
    简体中文 + 繁體中文. Every schema check passed — right row count,
    right keys, non-empty texts — so only comparing against the sheet
    reveals it. A wrong glossary is worse than none: it becomes the
    standard every later string is judged against."""
    from orbit8.exports import emit_bilingual_jsonl

    source = _glossary(tmp_path)
    out = tmp_path / "out.jsonl"

    # an adapter that ignored the instruction and took columns B and C
    def wrong_columns(paths, *, context="", **kw):
        return [("Barrier", "屏障", "屏障", str(paths[0])),
                ("Season Trial", "季节试炼", "季節試煉", str(paths[0]))]

    with pytest.raises(ValueError) as excinfo:
        emit_bilingual_jsonl([source], out, source_lang="en",
                             target_lang="ja",
                             columns=["English", "日本語"],
                             fallback=wrong_columns)
    assert "did not use the requested columns" in str(excinfo.value)
    assert not out.exists()          # nothing written


def test_the_right_columns_pass_the_check(tmp_path):
    from orbit8.exports import emit_bilingual_jsonl

    source = _glossary(tmp_path)

    def right_columns(paths, *, context="", **kw):
        assert "English" in context and "日本語" in context
        return [("Barrier", "Barrier", "バリア", str(paths[0])),
                ("Season Trial", "Season Trial", "季節の試練",
                 str(paths[0]))]

    written, _empty = emit_bilingual_jsonl(
        [source], tmp_path / "out.jsonl", source_lang="en",
        target_lang="ja", columns=["English", "日本語"],
        fallback=right_columns)
    assert written == 2


def test_a_column_name_that_does_not_exist_is_reported(tmp_path):
    from orbit8.exports import emit_bilingual_jsonl

    source = _glossary(tmp_path)
    with pytest.raises(ValueError) as excinfo:
        emit_bilingual_jsonl(
            [source], tmp_path / "out.jsonl", source_lang="en",
            target_lang="ja", columns=["English", "Klingon"],
            fallback=lambda p, *, context="", **kw: [
                ("A", "x", "y", str(p[0]))])
    assert "Klingon" in str(excinfo.value)


def _multi_locale_chat(tmp_path):
    from orbit8.llm import EchoProvider
    from orbit8.orchestrator import ChatOrchestrator
    job = Job.init(tmp_path / "proj" / "jobs", "j",
                   intake=IntakeBrief(game="G", source_lang="en",
                                      target_locales=["zh-CN", "zh-TW",
                                                      "ja", "ko"]),
                   source_files=[])
    return ChatOrchestrator(job, EchoProvider("ja"), operator="t",
                            dry_run=True)


def test_one_call_converts_every_locale(tmp_path):
    """Reading the same sheet once per locale is waste. The OUTPUT stays
    one file per locale because that is what the LQA cascade consumes —
    T1 checks width against ONE target, T2 consistency within ONE locale
    — but the input should be read once."""
    _glossary(tmp_path)
    result = json.loads(_multi_locale_chat(tmp_path)._t_standardize({
        "files": ["10-received/Glossary.xlsx"],
        "output": "bilingual_jsonl", "source_column": "English",
        "column_map": {"zh-CN": "简体中文", "zh-TW": "繁體中文",
                       "ja": "日本語", "ko": "한국어"},
        "out_name": "glossary"}))

    assert result["status"] == "complete"
    assert {row["locale"] for row in result["written"]} == {
        "zh-CN", "zh-TW", "ja", "ko"}
    assert all(row["written"] == 2 for row in result["written"])


def test_each_locale_gets_its_own_correct_column(tmp_path):
    """The failure this must not have: one adapter cached for `ja` being
    reused for `ko` and quietly emitting Japanese."""
    _glossary(tmp_path)
    chat = _multi_locale_chat(tmp_path)
    result = json.loads(chat._t_standardize({
        "files": ["10-received/Glossary.xlsx"],
        "output": "bilingual_jsonl", "source_column": "English",
        "column_map": {"ja": "日本語", "ko": "한국어"},
        "out_name": "glossary"}))

    by_locale = {}
    for row in result["written"]:
        first = json.loads(
            Path(row["path"]).read_text(encoding="utf-8").splitlines()[0])
        by_locale[row["locale"]] = first["target_text"]
    assert by_locale["ja"] == "バリア"
    assert by_locale["ko"] == "방벽"


def test_no_model_call_is_needed_for_a_term_sheet(tmp_path):
    """A named-column table is deterministic to read. Using a generated
    adapter would key the cache by suffix while the script hardcodes one
    language's column — the reuse bug above."""
    _glossary(tmp_path)
    chat = _multi_locale_chat(tmp_path)      # dry_run=True: no adapter
    result = json.loads(chat._t_standardize({
        "files": ["10-received/Glossary.xlsx"],
        "output": "bilingual_jsonl", "source_column": "English",
        "column_map": {"ja": "日本語"}, "out_name": "g"}))
    assert result["status"] == "complete"


def test_one_bad_locale_does_not_cost_the_others(tmp_path):
    """Per-locale failure: the operator needs to know which column to
    fix, and the good ones should still be written."""
    _glossary(tmp_path)
    result = json.loads(_multi_locale_chat(tmp_path)._t_standardize({
        "files": ["10-received/Glossary.xlsx"],
        "output": "bilingual_jsonl", "source_column": "English",
        "column_map": {"ja": "日本語", "ko": "Klingon"},
        "out_name": "glossary"}))

    assert result["status"] == "partial"
    assert [row["locale"] for row in result["written"]] == ["ja"]
    assert any("Klingon" in message for message in result["failed"])


def test_a_locale_outside_the_job_is_refused(tmp_path):
    _glossary(tmp_path)
    result = json.loads(_multi_locale_chat(tmp_path)._t_standardize({
        "files": ["10-received/Glossary.xlsx"],
        "output": "bilingual_jsonl", "source_column": "English",
        "column_map": {"fr": "Français"}, "out_name": "g"}))
    assert result["status"] == "failed"
    assert any("not a target locale" in m for m in result["failed"])


def test_column_map_needs_a_source_column(tmp_path):
    _glossary(tmp_path)
    result = _multi_locale_chat(tmp_path)._t_standardize({
        "files": ["10-received/Glossary.xlsx"],
        "output": "bilingual_jsonl",
        "column_map": {"ja": "日本語"}})
    assert "source_column" in result


def test_standardize_requires_exactly_two_columns(tmp_path):
    _glossary(tmp_path)
    result = _chat(tmp_path)._t_standardize({
        "files": ["10-received/Glossary.xlsx"],
        "output": "bilingual_jsonl", "target_lang": "ja",
        "columns": ["English"]})
    assert "exactly two names" in result


def test_standardize_can_write_outside_the_job_exports(tmp_path):
    """A glossary belongs in the PROJECT's 40-reference/glossary/, where
    every later stage resolves it — not in one job's exports/."""
    _glossary(tmp_path)
    chat = _chat(tmp_path)
    with pytest.raises(ValueError) as excinfo:      # dry-run stops later
        chat._t_standardize({
            "files": ["10-received/Glossary.xlsx"],
            "output": "bilingual_jsonl", "target_lang": "ja",
            "columns": ["English", "日本語"],
            "out_dir": "40-reference/glossary"})
    # reaching the adapter proves out_dir resolved inside the project
    assert "dry-run cannot generate an adapter" in str(excinfo.value)


def test_out_dir_cannot_escape_the_project(tmp_path):
    """out_dir is a real path from a model, so it goes through the WRITE
    confinement like every other one — the turn loop turns the refusal
    into an observation the agent sees."""
    from orbit8.tenancy import TenantError

    _glossary(tmp_path)
    with pytest.raises(TenantError):
        _chat(tmp_path)._t_standardize({
            "files": ["10-received/Glossary.xlsx"],
            "output": "source_json", "out_dir": "/etc"})


# ---------------------------------------------- the bilingual contract

def test_the_pair_validator_requires_all_three_fields():
    from orbit8.codegen import validate_pairs
    with pytest.raises(ValueError):
        validate_pairs('[{"key": "A", "text": "Start"}]', "f")


def test_the_pair_validator_keeps_empty_targets():
    """Opposite of the ingest validator, on purpose: for ingest an empty
    string is noise, for LQA it is the finding. Kept as long as SOMETHING
    was translated — see the next test for the all-empty case."""
    from orbit8.codegen import validate_pairs
    pairs = validate_pairs(
        '[{"key":"A","source":"Start","target":"開始"},'
        ' {"key":"B","source":"Quit","target":""}]', "f")
    assert ("B", "Quit", "", "f") in pairs


def test_an_all_empty_result_is_unsatisfiable_not_retryable():
    """Every target empty is a fact about the INPUT, not a defect in the
    generated code, so retrying spends model calls to fail identically.
    The observed case burned ~190s and four calls trying to pair a
    single-language file with itself."""
    from orbit8.codegen import UnsatisfiableInput, validate_pairs
    with pytest.raises(UnsatisfiableInput) as excinfo:
        validate_pairs('[{"key":"A","source":"ゲーム開始","target":""}]', "f")
    assert "ONE language" in str(excinfo.value)
    assert "BOTH" in str(excinfo.value)


def test_an_unsatisfiable_input_stops_the_retry_loop():
    """One attempt, not MAX_ATTEMPTS."""
    from orbit8.codegen import (BILINGUAL_SYSTEM, UnsatisfiableInput,
                                generate_converter, validate_pairs)

    calls = {"n": 0}

    class _OneLanguageProvider:
        name = "stub"
        model = "stub"

        def complete(self, system, prompt, **kw):
            calls["n"] += 1
            return ('import json,sys\n'
                    'print(json.dumps([{"key":"A","source":"x",'
                    '"target":""}]))')

    import tempfile
    path = Path(tempfile.mkstemp(suffix=".csv")[1])
    path.write_text("Key,Japanese\nA,ゲーム開始\n", encoding="utf-8")
    with pytest.raises(UnsatisfiableInput):
        generate_converter(_OneLanguageProvider(), path,
                           system_prompt=BILINGUAL_SYSTEM,
                           validate=validate_pairs)
    assert calls["n"] == 1                  # not 3


def test_the_pair_validator_drops_rows_with_no_source():
    from orbit8.codegen import validate_pairs
    with pytest.raises(ValueError):
        validate_pairs('[{"key":"A","source":"","target":"开始"}]', "f")


def test_the_pair_validator_refuses_duplicate_keys():
    from orbit8.codegen import validate_pairs
    with pytest.raises(ValueError):
        validate_pairs('[{"key":"A","source":"x","target":"y"},'
                       ' {"key":"A","source":"z","target":"w"}]', "f")


def test_a_stored_adapter_is_re_run_behind_the_same_wall():
    """Reuse must not become a way around validation: a stored bilingual
    adapter is re-validated against the pair contract, not the ingest
    one."""
    import inspect

    from orbit8.codegen import run_adapter
    assert "validate" in inspect.signature(run_adapter).parameters


# --------------------------------- the agent's own route into the cascade

def test_the_agent_can_audit_a_bilingual_jsonl(tmp_path):
    """`run_external_lqa` lived in the CLI and NOT in the tool set, so an
    agent that had just produced pairs_en-ja.jsonl with `standardize` had
    nowhere to take it — `scan_po` reads .po only. The conversion was a
    dead end, and the agent correctly reported it could not proceed."""
    chat = _multi_locale_chat(tmp_path)
    pairs = tmp_path / "proj" / "pairs.jsonl"
    pairs.write_text(json.dumps({
        "key": "A", "source_language": "en", "target_language": "ja",
        "source_text": "Start Game", "target_text": "ゲーム開始"},
        ensure_ascii=False) + "\n", encoding="utf-8")

    result = json.loads(chat._t_lqa_run(
        {"pairs": "pairs.jsonl", "locale": "ja",
         "deterministic_only": True}))
    assert result["status"] == "complete"
    assert result["checked"] == 1


def test_lqa_run_is_in_the_tool_set():
    from orbit8.orchestrator import ChatOrchestrator
    assert "lqa_run" in ChatOrchestrator.tool_names()


def test_each_locale_audit_gets_its_own_name(tmp_path):
    """One report per language pair — a shared name would overwrite every
    earlier locale's findings."""
    chat = _multi_locale_chat(tmp_path)
    pairs = tmp_path / "proj" / "p.jsonl"
    pairs.write_text(json.dumps({
        "key": "A", "source_language": "en", "target_language": "ja",
        "source_text": "Start", "target_text": "開始"},
        ensure_ascii=False) + "\n", encoding="utf-8")

    first = json.loads(chat._t_lqa_run(
        {"pairs": "p.jsonl", "locale": "ja", "deterministic_only": True}))
    second = json.loads(chat._t_lqa_run(
        {"pairs": "p.jsonl", "locale": "ko", "deterministic_only": True}))
    assert first["name"] != second["name"]


def test_an_unknown_locale_is_refused(tmp_path):
    chat = _multi_locale_chat(tmp_path)
    pairs = tmp_path / "proj" / "p.jsonl"
    pairs.write_text("{}\n", encoding="utf-8")
    assert "not a target locale" in chat._t_lqa_run(
        {"pairs": "p.jsonl", "locale": "fr"})


def test_a_missing_pairs_file_is_reported(tmp_path):
    chat = _multi_locale_chat(tmp_path)
    assert "no pairs file" in chat._t_lqa_run(
        {"pairs": "nope.jsonl", "locale": "ja"})


# ----------------------------------------- listing a directory usefully

def test_a_listing_names_the_directory_and_its_children(tmp_path):
    """A bare name list left the model unable to say what to pass next,
    so it re-listed the same directory five times before the loop breaker
    stopped it."""
    chat = _multi_locale_chat(tmp_path)
    received = tmp_path / "proj" / "10-received" / "drop"
    received.mkdir(parents=True)
    (received / "a.xlsx").write_bytes(b"x")

    result = json.loads(chat._t_list_files({"dir": "10-received"}))
    assert result["dir"].endswith("10-received")
    assert result["subdirectories"] == ["10-received/drop"]
    assert "not the same 'dir'" in result["note"]


def test_a_listing_without_subdirectories_has_no_note(tmp_path):
    chat = _multi_locale_chat(tmp_path)
    flat = tmp_path / "proj" / "flat"
    flat.mkdir(parents=True)
    (flat / "a.txt").write_text("x", encoding="utf-8")

    result = json.loads(chat._t_list_files({"dir": "flat"}))
    assert "subdirectories" not in result
    assert result["entries"] == ["a.txt (1B)"]


# ------------------------------- the glossary an external audit can see

def _term_sheet(tmp_path: Path) -> Path:
    pytest.importorskip("openpyxl")
    import openpyxl
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["English", "日本語"])
    sheet.append(["Barrier", "バリア"])
    sheet.append(["Season Trial", "季節の試練"])
    path = tmp_path / "terms.xlsx"
    book.save(path)
    return path


def test_a_term_sheet_becomes_the_t1_shape(tmp_path):
    """`standardize` produces bilingual PAIRS — the LQA input. The gate
    needs the T1 {metadata, terms} termbase, and nothing bridged the two,
    so a project could hold a perfectly good glossary and still audit with
    zero locked terms."""
    from orbit8.cli import main

    out = tmp_path / "glossary_terms.json"
    assert main(["glossary", "from-sheet", str(_term_sheet(tmp_path)),
                 "--source-column", "English", "--target-column", "日本語",
                 "--locale", "ja", "--out", str(out)]) == 0

    payload = json.loads(out.read_text(encoding="utf-8"))
    assert set(payload) == {"metadata", "terms"}
    assert payload["terms"]["Barrier"]["translation"] == "バリア"
    assert payload["metadata"]["locale"] == "ja"


def test_terms_are_advisory_until_ratified(tmp_path):
    """A client sheet is a proposal. Enforcing every row as law reports
    correct translations as defects — and the sheet may itself be wrong
    (one observed had a term whose Traditional column held an unrelated
    word)."""
    from orbit8.cli import main

    out = tmp_path / "g.json"
    main(["glossary", "from-sheet", str(_term_sheet(tmp_path)),
          "--source-column", "English", "--target-column", "日本語",
          "--locale", "ja", "--out", str(out)])
    assert not json.loads(out.read_text())["terms"]["Barrier"]["locked"]

    locked = tmp_path / "locked.json"
    main(["glossary", "from-sheet", str(_term_sheet(tmp_path)),
          "--source-column", "English", "--target-column", "日本語",
          "--locale", "ja", "--lock", "--out", str(locked)])
    assert json.loads(locked.read_text())["terms"]["Barrier"]["locked"]


def test_a_missing_column_is_reported(tmp_path, capsys):
    from orbit8.cli import main
    assert main(["glossary", "from-sheet", str(_term_sheet(tmp_path)),
                 "--source-column", "English", "--target-column", "Klingon",
                 "--locale", "ja", "--out", str(tmp_path / "g.json")]) == 2
    assert "Klingon" in capsys.readouterr().err


def test_an_external_audit_sees_the_promoted_termbase(tmp_path):
    """THE fix. `_glossary` read only the job's s3 artifact, which an
    external audit never has — it enters at S5 and never runs ASSET. So
    the terminology check ran with no locked terms and found nothing,
    which reads as a clean bill of health rather than a check that never
    ran."""
    from orbit8.cli import main

    project = tmp_path / "proj"
    (project / "20-work").mkdir(parents=True)
    (project / "40-reference" / "glossary").mkdir(parents=True)
    main(["glossary", "from-sheet", str(_term_sheet(tmp_path)),
          "--source-column", "English", "--target-column", "日本語",
          "--locale", "ja", "--lock",
          "--out", str(project / "40-reference" / "glossary"
                       / "glossary_terms.json")])

    job = Job.init(project / "jobs", "audit",
                   intake=IntakeBrief(game="G", source_lang="en",
                                      target_locales=["ja"]),
                   source_files=[])
    glossary = job._glossary("ja")
    assert glossary is not None
    assert glossary.locked_map(locked_only=True)["Barrier"] == "バリア"


def test_a_locked_term_violation_is_caught_end_to_end(tmp_path):
    """The whole point: sheet → termbase → audit → finding."""
    from orbit8.cli import main
    from orbit8.llm import EchoProvider
    from orbit8.orchestrator import ChatOrchestrator

    project = tmp_path / "proj"
    (project / "20-work").mkdir(parents=True)
    (project / "40-reference" / "glossary").mkdir(parents=True)
    main(["glossary", "from-sheet", str(_term_sheet(tmp_path)),
          "--source-column", "English", "--target-column", "日本語",
          "--locale", "ja", "--lock",
          "--out", str(project / "40-reference" / "glossary"
                       / "glossary_terms.json")])

    pairs = project / "20-work" / "pairs.jsonl"
    pairs.write_text("\n".join(json.dumps(row, ensure_ascii=False) for row in [
        {"key": "A", "source_language": "en", "target_language": "ja",
         "source_text": "Use the Barrier to block it",
         "target_text": "漁網で防いでください"},          # violates
        {"key": "B", "source_language": "en", "target_language": "ja",
         "source_text": "The Season Trial begins",
         "target_text": "季節の試練が始まる"},            # correct
    ]) + "\n", encoding="utf-8")

    job = Job.init(project / "jobs", "audit",
                   intake=IntakeBrief(game="G", source_lang="en",
                                      target_locales=["ja"]),
                   source_files=[])
    chat = ChatOrchestrator(job, EchoProvider("ja"), operator="t",
                            dry_run=True)
    result = json.loads(chat._t_lqa_run(
        {"pairs": str(pairs), "locale": "ja", "name": "t",
         "deterministic_only": True}))

    assert result["checked"] == 2
    assert result["by_bug_type"].get("terminology") == 1


# ------------------------------- the adapter cache must know its columns

def test_two_locales_do_not_share_one_cached_adapter(tmp_path):
    """The dead end: the cache key was suffix + file count while the
    GENERATED SCRIPT hardcodes one language's column. The adapter written
    for 日本語 was reused for 繁體中文, returned Japanese, the column guard
    correctly refused — and every retry hit the same cached script, so the
    request could never succeed."""
    from orbit8.exports import emit_bilingual_jsonl

    source = _glossary(tmp_path)
    seen = []

    def adapter(paths, *, context="", cache_key="", regenerate=False):
        seen.append(cache_key)
        # honour whichever columns the context asked for
        if "日本語" in context:
            return [("Barrier", "Barrier", "バリア", str(paths[0])),
                    ("Season Trial", "Season Trial", "季節の試練",
                     str(paths[0]))]
        return [("Barrier", "Barrier", "屏障", str(paths[0])),
                ("Season Trial", "Season Trial", "季節試煉", str(paths[0]))]

    emit_bilingual_jsonl([source], tmp_path / "ja.jsonl", source_lang="en",
                         target_lang="ja", columns=["English", "日本語"],
                         fallback=adapter)
    emit_bilingual_jsonl([source], tmp_path / "tw.jsonl", source_lang="en",
                         target_lang="zh-TW",
                         columns=["English", "繁體中文"], fallback=adapter)

    assert len(seen) == 2
    assert seen[0] != seen[1], "different columns must not share a cache key"


def test_a_stale_bad_adapter_is_regenerated_once(tmp_path):
    """A CACHED adapter can be wrong for these columns, and re-running it
    can only fail identically — before this the only recovery was deleting
    the artifact by hand."""
    from orbit8.exports import emit_bilingual_jsonl

    source = _glossary(tmp_path)
    calls = {"n": 0}

    def flaky(paths, *, context="", cache_key="", regenerate=False):
        calls["n"] += 1
        if not regenerate:                       # the stale cached script
            return [("Barrier", "屏障", "屏障", str(paths[0]))]
        return [("Barrier", "Barrier", "バリア", str(paths[0])),
                ("Season Trial", "Season Trial", "季節の試練",
                 str(paths[0]))]

    written, _empty = emit_bilingual_jsonl(
        [source], tmp_path / "out.jsonl", source_lang="en",
        target_lang="ja", columns=["English", "日本語"], fallback=flaky)
    assert written == 2
    assert calls["n"] == 2                       # once stale, once fresh


def test_a_second_failure_still_propagates(tmp_path):
    """Regenerating buys ONE retry. A genuinely unsatisfiable request must
    still surface rather than loop."""
    from orbit8.exports import emit_bilingual_jsonl

    source = _glossary(tmp_path)
    with pytest.raises(ValueError) as excinfo:
        emit_bilingual_jsonl(
            [source], tmp_path / "out.jsonl", source_lang="en",
            target_lang="ja", columns=["English", "日本語"],
            fallback=lambda p, *, context="", cache_key="",
            regenerate=False: [("Barrier", "屏障", "屏障", str(p[0]))])
    assert "did not use the requested columns" in str(excinfo.value)

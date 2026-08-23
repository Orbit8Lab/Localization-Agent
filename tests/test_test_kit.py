"""In-game test kit emitter (docs/STANDARDS.md §4.5)."""
import pytest

from orbit8 import standards as st
from orbit8.test_kit import emit_test_kit, normalize_error_types


def _rows():
    return [
        {"StringID": "UI_0001", "StringType": "UI",
         "StringLocation": "/Game/UI/HUD", "Source": "Attack",
         "Target_Current": "攻击", "AI_ExpectedBug": "No",
         "AI_ErrorTypes": ""},
        {"StringID": "SKL_0042", "StringType": "Skill",
         "StringLocation": "/Game/Skills", "Source": "Rally the troops",
         "Target_Current": "集结部队", "AI_ExpectedBug": "Yes",
         "AI_ErrorTypes": ["length", "terminology"]},
        {"StringID": "DLG_0100", "StringType": "Dialogue",
         "StringLocation": "/Game/Dialog/Ch1", "Source": "Hello {0}!",
         "Target_Current": "你好 {0}！", "AI_ExpectedBug": "Yes",
         "AI_ErrorTypes": "Tag/Markup"},
    ]


def _sheet(path):
    import openpyxl
    return openpyxl.load_workbook(path)["Test Kit"]


def test_emits_standard_headers(tmp_path):
    out = tmp_path / "test_kit_zh-CN.xlsx"
    assert emit_test_kit(out, _rows()) == 3
    header = [c.value for c in _sheet(out)[1]]
    assert header == list(st.TEST_KIT_FORM_HEADERS)


def test_tester_columns_emit_empty(tmp_path):
    """The tester owns TEST_*; a value passed in for one is ignored."""
    out = tmp_path / "kit.xlsx"
    rows = _rows()
    rows[0]["TEST_Decision"] = "No Issue"      # must NOT survive
    rows[0]["TEST_Notes"] = "prefilled"
    emit_test_kit(out, rows)
    sheet = _sheet(out)
    header = [c.value for c in sheet[1]]
    for col in st.TEST_FILLED_COLUMNS:
        column = header.index(col) + 1
        for row in range(2, sheet.max_row + 1):
            assert sheet.cell(row=row, column=column).value in (None, "")


def test_every_string_ships_regardless_of_prediction(tmp_path):
    """A kit that dropped AI_ExpectedBug=No rows could never catch a missed
    prediction — the low-risk rows must still be present."""
    out = tmp_path / "kit.xlsx"
    emit_test_kit(out, _rows())
    sheet = _sheet(out)
    ids = {sheet.cell(row=r, column=1).value
           for r in range(2, sheet.max_row + 1)}
    assert ids == {"UI_0001", "SKL_0042", "DLG_0100"}


def test_risk_first_ordering_is_stable(tmp_path):
    out = tmp_path / "kit.xlsx"
    emit_test_kit(out, _rows())
    sheet = _sheet(out)
    ids = [sheet.cell(row=r, column=1).value
           for r in range(2, sheet.max_row + 1)]
    # Yes rows first, original relative order kept within each group
    assert ids == ["SKL_0042", "DLG_0100", "UI_0001"]


def test_risk_first_can_be_disabled(tmp_path):
    out = tmp_path / "kit.xlsx"
    emit_test_kit(out, _rows(), risk_first=False)
    sheet = _sheet(out)
    ids = [sheet.cell(row=r, column=1).value
           for r in range(2, sheet.max_row + 1)]
    assert ids == ["UI_0001", "SKL_0042", "DLG_0100"]


def test_error_types_normalize_from_tokens():
    assert normalize_error_types(["length", "terminology"]) == (
        "Length/Truncation; Terminology")
    assert normalize_error_types(["Tag/Markup"]) == "Tag/Markup"
    assert normalize_error_types([]) == ""
    assert normalize_error_types(None) == ""


def test_error_types_dedup_preserving_order():
    assert normalize_error_types(
        ["placeholder", "markup", "terminology"]) == "Tag/Markup; Terminology"


def test_unknown_error_type_is_refused():
    """An unaggregatable label is a caller bug, caught at write time."""
    with pytest.raises(ValueError, match="unknown predicted error type"):
        normalize_error_types(["overflow"])


def test_list_error_types_render_into_the_cell(tmp_path):
    out = tmp_path / "kit.xlsx"
    emit_test_kit(out, _rows(), risk_first=False)
    sheet = _sheet(out)
    column = [c.value for c in sheet[1]].index("AI_ErrorTypes") + 1
    assert sheet.cell(row=3, column=column).value == (
        "Length/Truncation; Terminology")


def test_dropdowns_attached_to_tester_columns(tmp_path):
    out = tmp_path / "kit.xlsx"
    emit_test_kit(out, _rows())
    sheet = _sheet(out)
    header = [c.value for c in sheet[1]]
    validated = set()
    for validation in sheet.data_validations.dataValidation:
        letters = {str(rng).split("!")[-1].split("$")[-1].rstrip("0123456789")
                   for rng in validation.sqref.ranges}
        for cell_range in validation.sqref.ranges:
            col = cell_range.min_col
            validated.add(header[col - 1])
    assert {"TEST_Decision", "TEST_AssignedTeam", "AI_ExpectedBug",
            "StringType"} <= validated


def test_extra_columns_extend_but_never_alter(tmp_path):
    out = tmp_path / "kit.xlsx"
    rows = _rows()
    rows[0]["Findings"] = "gate: length"
    emit_test_kit(out, rows, extra_columns=("Findings",))
    header = [c.value for c in _sheet(out)[1]]
    assert header[:len(st.TEST_KIT_FORM_HEADERS)] == list(
        st.TEST_KIT_FORM_HEADERS)
    assert header[-1] == "Findings"


def test_extra_column_collision_is_refused(tmp_path):
    with pytest.raises(ValueError, match="collides"):
        emit_test_kit(tmp_path / "kit.xlsx", _rows(),
                      extra_columns=("Source",))


def test_guide_sheet_present(tmp_path):
    import openpyxl
    out = tmp_path / "kit.xlsx"
    emit_test_kit(out, _rows())
    book = openpyxl.load_workbook(out)
    assert "说明 How to fill" in book.sheetnames
    text = "\n".join(str(row[0].value) for row in book["说明 How to fill"].rows)
    # the two rules a tester most easily gets wrong
    assert "does NOT mean the string is fine" in text
    assert "Blank = not tested" in text


def test_empty_kit_is_valid(tmp_path):
    out = tmp_path / "kit.xlsx"
    assert emit_test_kit(out, []) == 0
    assert [c.value for c in _sheet(out)[1]] == list(st.TEST_KIT_FORM_HEADERS)

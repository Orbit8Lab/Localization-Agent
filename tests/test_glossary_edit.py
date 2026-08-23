"""Direct operator edits to a T1 glossary."""
import json
from pathlib import Path

import openpyxl

from orbit8.glossary_edit import (TermEdit, apply_edits,
                                  edit_glossary_file, parse_term_arg)

BASE = {
    "metadata": {"game": "测试", "locale": "en",
                 "family_rules": {"致幻": {"translation":
                                           "Hallucination"}}},
    "terms": {
        "车队": {"translation": "Convoy", "type": "decision",
                 "locked": True, "evidence": "M-04"},
        "马车": {"translation": "Wagon", "type": "mined-llm",
                 "locked": False, "evidence": "corpus ×32",
                 "category": "Item"},
        "狼人": {"translation": "Werewolf", "type": "mined",
                 "locked": False, "evidence": "corpus ×23"},
        "致幻陷阱": {"translation": "Hallucinogenic Trap",
                     "type": "mined", "locked": False,
                     "evidence": "corpus ×2"},
    }}


def _fresh() -> dict:
    return json.loads(json.dumps(BASE))


def test_parse_term_arg_forms():
    assert parse_term_arg("任务=Mission").zh == "任务"
    multi = parse_term_arg("马车/车队=Convoy")
    assert multi.zh == "马车/车队" and multi.en == "Convoy"
    rename = parse_term_arg("狼人>冥界使徒=Underworld Apostle")
    assert rename.zh == "冥界使徒" and rename.zh_old == "狼人"


def test_add_new_term_locks_it():
    g = _fresh()
    result = apply_edits(g, [TermEdit(zh="BOSS", en="Boss")],
                         origin="operator 2026-08-03")
    entry = g["terms"]["BOSS"]
    assert entry["locked"] and entry["type"] == "operator"
    assert entry["evidence"] == "operator 2026-08-03"
    assert result.added == [{"zh": "BOSS", "en": "Boss"}]
    assert list(g["terms"])[0] in ("BOSS", "车队")     # locked sort first


def test_overwrites_unlocked_records_previous():
    g = _fresh()
    result = apply_edits(g, [TermEdit(zh="马车", en="Convoy")],
                         origin="op")
    entry = g["terms"]["马车"]
    assert entry["locked"] and entry["translation"] == "Convoy"
    assert entry["supersedes_rendering"] == "Wagon"
    assert entry["category"] == "Item"          # metadata carried
    assert result.overwritten[0]["was"] == "Wagon"


def test_locked_conflict_needs_force():
    g = _fresh()
    result = apply_edits(g, [TermEdit(zh="车队", en="Caravan")],
                         origin="op")
    assert g["terms"]["车队"]["translation"] == "Convoy"   # untouched
    assert result.conflicts[0]["current"] == "Convoy"
    assert not result.wrote

    g2 = _fresh()
    forced = apply_edits(g2, [TermEdit(zh="车队", en="Caravan")],
                         origin="op", force=True)
    assert g2["terms"]["车队"]["translation"] == "Caravan"
    assert forced.overwritten and forced.wrote


def test_identical_locked_edit_is_noop():
    g = _fresh()
    result = apply_edits(g, [TermEdit(zh="车队", en="Convoy")],
                         origin="op")
    assert result.unchanged and not result.conflicts and not result.wrote


def test_aliases_and_rename_retire():
    g = _fresh()
    result = apply_edits(
        g, [TermEdit(zh="冥界使徒", en="Underworld Apostle",
                     zh_old="狼人")], origin="op")
    assert g["terms"]["冥界使徒"]["supersedes"] == "狼人"
    assert "狼人" not in g["terms"]                    # retired
    assert result.retired[0]["was"] == "Werewolf"

    # 马车/车队=Convoy: 马车 (unlocked Wagon) is overwritten and becomes
    # canonical; 车队 already rules Convoy → left untouched as unchanged,
    # never re-stamped just to add an alias marker
    g2 = _fresh()
    multi = apply_edits(g2, [TermEdit(zh="马车/车队", en="Convoy")],
                        origin="op")
    assert g2["terms"]["马车"]["translation"] == "Convoy"
    assert g2["terms"]["马车"]["supersedes_rendering"] == "Wagon"
    assert [u["zh"] for u in multi.unchanged] == ["车队"]

    # a genuinely new alias does get the marker
    g3 = _fresh()
    apply_edits(g3, [TermEdit(zh="任务/委托", en="Mission")], origin="op")
    assert g3["terms"]["委托"]["alias_of"] == "任务"


def test_family_rule_recheck_and_file_roundtrip(tmp_path: Path):
    path = tmp_path / "glossary_terms.json"
    path.write_text(json.dumps(BASE, ensure_ascii=False),
                    encoding="utf-8")
    glossary, result, backup = edit_glossary_file(
        path, [TermEdit(zh="任务", en="Mission")], origin="op",
        backup_stamp="20260803")
    # backup kept, json + xlsx both refreshed
    assert backup and backup.exists()
    assert json.loads(path.read_text())["terms"]["任务"]["locked"]
    book = openpyxl.load_workbook(tmp_path / "glossary_terms.xlsx")
    zh_col = [r[0] for r in book["Glossary"].iter_rows(min_row=2,
                                                       values_only=True)]
    assert "任务" in zh_col
    assert "Family Rules 术语族" in book.sheetnames
    # unlocked member still violates the 致幻 family rule → flagged
    assert any(f["zh"] == "致幻陷阱" for f in result.flagged)


def test_unify_declares_a_source_variant_group():
    """Messy source: 马车 and 车队 are the same thing; both must render
    as Convoy, and the variant must keep its entry so the LQA gate still
    matches those source strings."""
    from orbit8.glossary_check import check_glossary
    from orbit8.glossary_edit import unify_terms
    g = _fresh()
    summary = unify_terms(g, "车队", ["马车"], origin="2026-08-05")
    assert summary["variants"] == ["马车"]
    assert g["terms"]["车队"]["locked"] and \
        g["terms"]["车队"]["translation"] == "Convoy"
    variant = g["terms"]["马车"]
    assert variant["translation"] == "Convoy"      # normalized
    assert variant["variant_of"] == "车队"          # and declared
    assert variant["locked"]                        # gate-enforced
    # the checker now sees a declared group, not a contradiction
    report = check_glossary(g)
    assert not report.blocked
    assert any(i.kind == "source_variants" for i in report.issues)


def test_unify_can_retire_an_obsolete_spelling():
    from orbit8.glossary_edit import unify_terms
    g = _fresh()
    summary = unify_terms(g, "车队", ["马车"], origin="x", keep=False)
    assert summary["retired"] == ["马车"]
    assert "马车" not in g["terms"]
    assert g["terms"]["车队"]["supersedes"] == "马车"


def test_unify_reports_unknown_variants_and_needs_a_head():
    import pytest
    from orbit8.glossary_edit import unify_terms
    g = _fresh()
    summary = unify_terms(g, "车队", ["马车", "运输队"], origin="x")
    assert summary["missing"] == ["运输队"]
    with pytest.raises(KeyError):
        unify_terms(g, "全新词", ["马车"], origin="x")
    # …unless a translation is supplied to create it
    out = unify_terms(g, "全新词", ["马车"], translation="Brand New",
                      origin="x")
    assert out["translation"] == "Brand New"

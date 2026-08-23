"""Bug-report builder per the bug-report-builder skill: schema, sorting,
code-token blanks-with-comments, suggestion generation, and round-trip
merge-back of human columns."""
import json
from pathlib import Path

import openpyxl

from orbit8.bug_report import (CODE_TOKEN_NOTE, HEADERS, build_suggestions,
                               is_code_token, write_bug_report_xlsx,
                               write_tech_summary)
from orbit8.schemas import (BugType, Finding, LQAItem, LQAReport, Severity,
                            VerifiedFinding)


def _report():
    def item(uid, key, source, target, bug, sev):
        return LQAItem(uid=uid, game_keys=[key], source=source,
                       target=target,
                       findings=[VerifiedFinding(finding=Finding(
                           key=uid, bug_type=bug, severity=sev,
                           message=f"msg {uid}", evidence="e"))])
    return LQAReport(
        job_id="j", locale="zh-CN", checked=10, flagged_strings=3,
        findings_total=3, confirmed=3, overturned=0, uncertain=0,
        block_ship=True, by_severity={"high": 2, "medium": 1},
        by_bug_type={"terminology": 2, "untranslated": 1},
        items=[
            item("u1", "K1", "Journal", "日志", BugType.TERMINOLOGY,
                 Severity.HIGH),
            item("u2", "K2", "mogo69", "mogo69", BugType.UNTRANSLATED,
                 Severity.HIGH),
            item("u3", "K3", "Long settings text here", "很长的设置文本",
                 BugType.LENGTH, Severity.MEDIUM),
        ])


def test_code_token_heuristic():
    assert is_code_token("mogo69") and is_code_token("Loading")
    assert is_code_token("123") and is_code_token("%")
    assert not is_code_token("Press any key")
    assert not is_code_token("老国王低声说")


def test_xlsx_schema_sorting_and_code_token_comment(tmp_path: Path):
    path = tmp_path / "r.xlsx"
    count = write_bug_report_xlsx(_report(), path,
                                  suggestions={"u1": "手记"}, game="G")
    assert count == 3
    rows = list(openpyxl.load_workbook(path).active.iter_rows(
        values_only=True))
    assert list(rows[0]) == HEADERS
    severities = [r[HEADERS.index("Severity")] for r in rows[1:]]
    assert severities == ["HIGH", "HIGH", "MEDIUM"]      # sorted desc
    by_loc = {r[HEADERS.index("Location/String ID")]: r for r in rows[1:]}
    u1 = by_loc["u1 :: K1"]
    assert u1[HEADERS.index(
        "Expected Result / Suggested Translation")] == "手记"
    assert "[terminology]" in u1[HEADERS.index("Description")]
    u2 = by_loc["u2 :: K2"]                    # code token: blank + comment
    assert not u2[HEADERS.index("Expected Result / Suggested Translation")]
    assert u2[HEADERS.index("Orbit8 Comment")] == CODE_TOKEN_NOTE


def test_locations_map_renders_widget_paths(tmp_path: Path):
    path = tmp_path / "r.xlsx"
    write_bug_report_xlsx(
        _report(), path, suggestions={}, game="G",
        locations={"K1": "/Game/UI/WDG_Journal.WDG_Journal_C:"
                         "WidgetTree.TextBlock_0.Text"})
    rows = list(openpyxl.load_workbook(path).active.iter_rows(
        values_only=True))
    ids = [r[HEADERS.index("Location/String ID")] for r in rows[1:]]
    assert ("u1 :: /Game/UI/WDG_Journal.WDG_Journal_C:"
            "WidgetTree.TextBlock_0.Text :: K1") in ids
    assert "u2 :: K2" in ids                   # unknown key: bare fallback


def test_round_trip_preserves_human_columns(tmp_path: Path):
    path = tmp_path / "r.xlsx"
    write_bug_report_xlsx(_report(), path, suggestions={}, game="G")
    book = openpyxl.load_workbook(path)
    sheet = book.active
    for row in sheet.iter_rows(min_row=2):
        if row[HEADERS.index("Location/String ID")].value == "u1 :: K1":
            row[HEADERS.index("Status")].value = "Fixed"
            row[HEADERS.index("Dev Feedback")].value = "done in build 42"
            row[HEADERS.index("Orbit8 Comment")].value = "verified in game"
    book.save(path)

    write_bug_report_xlsx(_report(), path, suggestions={"u1": "手记"},
                          game="G")                       # rebuild
    assert path.with_suffix(".bak.xlsx").exists()          # prior backed up
    rows = list(openpyxl.load_workbook(path).active.iter_rows(
        values_only=True))
    by_loc = {r[HEADERS.index("Location/String ID")]: r for r in rows[1:]}
    u1 = by_loc["u1 :: K1"]
    assert u1[HEADERS.index("Status")] == "Fixed"          # human kept
    assert u1[HEADERS.index("Dev Feedback")] == "done in build 42"
    assert u1[HEADERS.index("Orbit8 Comment")] == "verified in game"
    u2 = by_loc["u2 :: K2"]                    # AUTO note refreshed, kept
    assert u2[HEADERS.index("Orbit8 Comment")] == CODE_TOKEN_NOTE


def test_suggestions_skip_code_tokens():
    class RepairProvider:
        name, model, tokens_spent = "s", "t", 0.0
        def complete(self, system, user, *, temperature=0.3,
                     max_tokens=2000):
            import re
            keys = [k for k in re.findall(r"^### (\S+)$", user, flags=re.M)
                    if k != "END"]
            return json.dumps({"items": [
                {"key": k, "target_text": f"fix-{k}", "term_decisions": {},
                 "notes": None} for k in keys]})
    suggestions = build_suggestions(RepairProvider(), _report().items,
                                    game="G", source_lang="en",
                                    locale="zh-CN")
    assert suggestions == {"u1": "fix-u1", "u3": "fix-u3"}   # u2 skipped


def test_tech_summary(tmp_path: Path):
    path = tmp_path / "s.md"
    write_tech_summary(_report(), path, game="G",
                       split_counts={"story": 3, "string": 7},
                       suggestions_count=2)
    text = path.read_text(encoding="utf-8")
    assert "BLOCK SHIP" in text and "story 3" in text
    assert "Localization - Terminology" in text
